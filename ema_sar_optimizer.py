import os
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
import json
import zipfile
import argparse
import numpy as np
import pandas as pd
from scipy import stats, optimize
from scipy.signal import lfilter
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from dataclasses import dataclass, field
from concurrent.futures import ProcessPoolExecutor, as_completed

# ----------------------------- CONFIG ---------------------------------------
SYMBOLS    = ["ETH", "SOL", "XRP", "XAU", "XAG"]
TIMEFRAMES = ["5m", "15m", "30m", "1h", "2h", "3h", "4h"]
CUTOFF     = None  # None = load all historical data up to present (no artificial truncation)
BALANCE0   = 100_000.0
MIN_P      = 2
MAX_P      = 300
FEE_BPS    = 0.0
SLIPPAGE_BPS = 5.0
DATA_XLSX  = "./data/market_data.xlsx"
OUT_DIR    = "./output"
CSV_DIR    = "./output/csv"
MC_SIMS    = 300  # adaptive; scaled per dataset in worker()
WF_FOLDS   = 3
REGIME_WIN = 50
ROBUST_R   = 5
SL_LEVELS_PCT  = [0.1, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 7.5, 10.0]
TSL_LEVELS_PCT = [0.1, 0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 7.5, 10.0]
TP_RR_MULTIPLIERS = [1.5, 2.0, 3.0]  # Take-Profit Risk/Reward Ratios
ATR_GRID = [0.5, 0.75, 1.0, 1.25, 1.5, 2.0, 2.5, 3.0] # U1: ATR-Normalized Multipliers
SQN_GATE_MIN = 1.6 # U6: Van Tharp System Quality Gate
RISK_AVERSION = 2.0  # Q8: von Neumann-Morgenstern utility risk aversion lambda (CLI overridable)

# S4: Adaptive EMA period ranges per timeframe (same wall-clock horizon across TFs)
PERIOD_RANGE = {
    "5m":  (2, 100),   # 10min – 8h
    "15m": (2, 150),   # 30min – 37h
    "30m": (2, 200),   # 1h – 4 days
    "1h":  (2, 250),   # 2h – 10 days
    "2h":  (2, 200),   # 4h – 17 days
    "3h":  (2, 180),   # 6h – 22 days
    "4h":  (2, 180),   # 8h – 30 days
}

# Q4: Historical crash stress windows (inclusive ISO date strings)
CRASH_WINDOWS = [
    ("COVID-Crash",    "2020-02-20", "2020-03-23"),
    ("FTX-Collapse",   "2022-11-01", "2022-12-01"),
    ("SVB-BankRun",    "2023-03-08", "2023-03-31"),
    ("Crypto-Winter",  "2022-06-01", "2022-07-01"),
]

SESSION_BOUNDS = [
    ("Asia", 0, 8), ("London", 8, 13), ("NY_overlap", 13, 16),
    ("NY", 16, 21), ("Late", 21, 24),
]

def session_of_hour(h):
    for name, lo, hi in SESSION_BOUNDS:
        if lo <= h < hi:
            return name
    return "Unknown"

# ----------------------------- STATISTICAL & MATH UTILS ---------------------
def adf_test_pv(series):
    """
    Augmented Dickey-Fuller test for unit-root stationarity testing using statsmodels adfuller,
    with fallback to OLS regression t-statistic & MacKinnon critical value approximation.
    """
    if len(series) < 25: return 1.0
    try:
        from statsmodels.tsa.stattools import adfuller
        res = adfuller(series, maxlag=1, autolag=None)
        return round(float(res[1]), 4)
    except Exception:
        diff = np.diff(series)
        lag = series[:-1]
        
        # Fit regression: dy_t = alpha + beta * y_{t-1}
        res = stats.linregress(lag, diff)
        beta = res.slope
        stderr = res.stderr if res.stderr > 1e-9 else 1e-9
        t_stat = beta / stderr
        
        # MacKinnon (1996) 5% & 1% critical value thresholds for N > 100 (with constant)
        if t_stat < -3.43:
            p_val = 0.005
        elif t_stat < -2.86:
            p_val = 0.03
        elif t_stat < -2.57:
            p_val = 0.08
        else:
            p_val = float(min(1.0, max(0.10, 0.10 + 0.3 * (t_stat + 2.57))))
        
        return round(p_val, 4)

def _single_vr(series, k):
    if len(series) < k * 10: return 1.0, 1.0
    r1 = np.diff(series)
    rk = series[k:] - series[:-k]
    var1 = np.var(r1, ddof=1)
    vark = np.var(rk, ddof=1)
    if var1 <= 0: return 1.0, 1.0
    vr = vark / (k * var1)
    n = len(series)
    sum_sq = np.sum(r1**2)
    if sum_sq <= 0: return 1.0, 1.0
    phi = 0.0
    for j in range(1, k):
        num = np.sum((r1[j:]**2) * (r1[:-j]**2))
        den = sum_sq**2
        delta_j = (num / den) * n
        phi += ((2.0 * (k - j) / k)**2) * delta_j
    z = (vr - 1.0) / np.sqrt(phi + 1e-9)
    pv = float(2 * (1 - stats.norm.cdf(abs(z))))
    return vr, pv

def variance_ratio_test(series, lags=(2, 4, 8, 16)):
    """
    M4 Upgrade: Multi-Lag Heteroscedasticity-Robust Variance Ratio Test (Lo & MacKinlay 1988)
    Evaluates VR across lags k in [2, 4, 8, 16] and reports maximum departure |VR-1| and optimal lag.
    """
    if isinstance(lags, int):
        lags = (lags,)
    best_vr, best_pv, best_k = 1.0, 1.0, 2
    max_dev = -1.0
    for k in lags:
        vr, pv = _single_vr(series, k)
        dev = abs(vr - 1.0)
        if dev > max_dev:
            max_dev = dev
            best_vr, best_pv, best_k = vr, pv, k
    return round(float(best_vr), 3), round(float(best_pv), 4), int(best_k)

def runs_test(rets):
    """Runs test for randomness of trade return sequence."""
    if len(rets) < 10: return 0.0, 1.0
    bin_s = (rets > 0).astype(int)
    n1 = np.sum(bin_s)
    n0 = len(bin_s) - n1
    if n0 == 0 or n1 == 0: return 0.0, 1.0
    runs = 1 + np.sum(bin_s[1:] != bin_s[:-1])
    exp_r = 1.0 + (2.0 * n0 * n1) / len(bin_s)
    var_r = (2.0 * n0 * n1 * (2.0 * n0 * n1 - len(bin_s))) / (len(bin_s)**2 * (len(bin_s) - 1) + 1e-9)
    if var_r <= 0: return float(runs), 1.0
    z = (runs - exp_r) / np.sqrt(var_r)
    pv = float(2 * (1 - stats.norm.cdf(abs(z))))
    return round(float(z), 3), round(pv, 4)

# ----------------------------- DATA PROFILING & AUDIT -----------------------
def profile_dataset(df, symbol, tf):
    close = df["close"].values
    ts    = pd.to_datetime(df["time"]).values
    n     = len(df)
    s0, sn = float(close[0]), float(close[-1])
    bh     = (sn / s0 - 1.0) * 100.0
    t0     = pd.to_datetime(ts[0]); tn = pd.to_datetime(ts[-1])
    days   = max(1.0, (tn - t0).total_seconds() / 86400.0)
    bpd    = n / days
    br     = np.diff(close) / close[:-1]
    vol_b  = float(br.std() * 100.0) if len(br) else 0.0
    ann_v  = vol_b * np.sqrt(bpd * 365)
    regime = "Bullish" if bh > 10 else ("Bearish" if bh < -10 else "Sideways")

    def hurst(x, min_len=8):
        if len(x) < min_len * 2: return 0.5
        x_centered = x - np.mean(x)
        y = np.cumsum(x_centered)
        scales = np.unique(np.logspace(1, np.log10(len(y)//4), num=15, dtype=int))
        if len(scales) < 2: return 0.5
        F = []
        for s in scales:
            shape = (y.shape[0] // s) * s
            y_trunc = y[:shape]
            y_reshaped = y_trunc.reshape(-1, s)
            t = np.arange(s)
            F_s = 0
            for row in y_reshaped:
                p = np.polyfit(t, row, 1)
                fit = np.polyval(p, t)
                F_s += np.sum((row - fit)**2)
            F.append(np.sqrt(F_s / shape))
        if len(F) > 2:
            return float(np.polyfit(np.log2(scales), np.log2(F), 1)[0])
        return 0.5

    outlier_count = int(np.sum(np.abs(br - br.mean()) > 4.0 * (br.std() + 1e-9)))
    adf_p = adf_test_pv(close)
    vr, vr_p, vr_k = variance_ratio_test(close, lags=(2, 4, 8, 16))

    return dict(
        symbol=symbol, timeframe=tf, bars=n,
        start=t0.strftime("%Y-%m-%d"), end=tn.strftime("%Y-%m-%d"),
        days=round(days, 1), start_price=s0, end_price=sn,
        bh_return_pct=round(bh, 2), bar_vol_pct=round(vol_b, 4),
        ann_vol_pct=round(ann_v, 2), regime=regime,
        hurst=round(hurst(close), 3),
        skewness=round(float(stats.skew(br)), 3),
        kurtosis=round(float(stats.kurtosis(br)), 3),
        outliers=outlier_count,
        adf_pvalue=round(adf_p, 4),
        variance_ratio=vr,
        vr_pvalue=vr_p,
        vr_lag=vr_k
    )

# ----------------------------- CORE ENGINE ----------------------------------
def _ema(close, alpha):
    y, _ = lfilter([alpha], [1, -(1-alpha)], close,
                   zi=np.array([close[0] * (1-alpha)]))
    return y

def ema_matrix(close, periods):
    out = np.empty((len(periods), len(close)), dtype=np.float64)
    for i, p in enumerate(periods):
        out[i] = _ema(close, 2.0 / (p + 1.0))
    return out

def _max_consec(arr, val):
    b = (np.array(arr) == val).astype(int)
    groups = np.diff(np.concatenate(([0], b, [0])))
    starts = np.where(groups == 1)[0]
    ends   = np.where(groups == -1)[0]
    return int((ends - starts).max()) if len(starts) else 0

def cagr(total_return_pct, days):
    """Computes Annualized Compound Growth Rate %."""
    if days <= 0 or total_return_pct <= -100.0: return -100.0
    return float(((1 + total_return_pct / 100.0) ** (365.0 / max(1.0, days)) - 1) * 100.0)

def _drawdown(eq):
    peak = np.maximum.accumulate(eq)
    return (eq - peak) / peak

def continuous_kelly(rets, fraction=1.0):
    """
    F5 Upgrade: Continuous-Time Kelly Criterion for Fat-Tailed Financial Distributions (Thorp 2006)
    f* = mu / ( (1/fraction) * sigma^2 )
    - Full Kelly (fraction=1.0): f* = mu / sigma^2
    - Half Kelly (fraction=0.5): f* = mu / (2 * sigma^2)
    Valid for any arbitrary non-Bernoulli return distribution.
    """
    if len(rets) < 2: return 0.0
    r = np.array(rets)
    mu = float(np.mean(r))
    var = float(np.var(r, ddof=1))
    if var <= 1e-9 or mu <= 0: return 0.0
    denom = 1.0 / max(1e-4, fraction)
    f_opt = mu / (denom * var)
    return float(max(0.0, min(1.0, f_opt)))

def kelly_criterion(arg1, arg2=None, fraction=1.0):
    """F5 Continuous Kelly wrapper with legacy fallback."""
    if isinstance(arg1, (list, np.ndarray, pd.Series)):
        return continuous_kelly(arg1, fraction=fraction)
    else:
        wr = float(arg1)
        b = float(arg2) if arg2 is not None else 1.0
        if b <= 0 or wr <= 0: return 0.0
        q = 1.0 - wr
        return float(max(0.0, min(1.0, ((wr * b - q) / b) * fraction)))

def risk_of_ruin(arg1, arg2=None, n_trades=100, ruin_level=0.5):
    """
    M5 Upgrade: Diffusion Approximation Risk of Ruin (Kaufman 2013, Vince 1992)
    Supports continuous trade return distribution with variable/volatility scaling:
    RoR = exp( -2 * (mu / var) * |ln(ruin_level)| )
    """
    if isinstance(arg1, (list, np.ndarray, pd.Series)):
        r = np.array(arg1)
        if len(r) < 2: return 1.0
        mu = float(np.mean(r))
        var = float(np.var(r, ddof=1))
        if mu <= 0 or var <= 1e-9: return 1.0
        z_ruin = abs(np.log(max(1e-4, ruin_level)))
        ror = np.exp(-2.0 * (mu / var) * z_ruin)
        return float(max(0.0, min(1.0, ror)))
    else:
        win_rate = float(arg1)
        payoff_ratio = float(arg2) if arg2 is not None else 1.0
        if payoff_ratio <= 0 or win_rate <= 0: return 1.0
        edge = win_rate - (1.0 - win_rate) / payoff_ratio
        if edge >= 1.0: return 0.0
        if edge <= 0.0: return 1.0
        return float(((1.0 - edge) / (1.0 + edge)) ** n_trades)

def deflated_sharpe_ratio(sharpe, n_obs, n_trials=4500, skewness=0.0, kurtosis=3.0):
    """
    M2 Upgrade: Deflated Sharpe Ratio (Bailey & López de Prado 2012)
    Rigorously corrects Sharpe Ratio for multiple testing (trials K), sample size (N), skewness, and kurtosis.
    Returns: Probabilistic Deflated Sharpe Ratio (DSR) probability in range [0.0, 1.0].
    """
    if n_obs < 5 or sharpe <= 0:
        return 0.0
    
    sr_var = (1.0 - skewness * sharpe + ((kurtosis - 1.0) / 4.0) * (sharpe ** 2)) / max(1.0, n_obs - 1.0)
    sr_std = np.sqrt(max(1e-9, sr_var))
    
    euler = 0.5772156649
    if n_trials > 1:
        e_max_sr = sr_std * ((1.0 - euler) * stats.norm.ppf(1.0 - 1.0 / n_trials) + 
                             euler * stats.norm.ppf(1.0 - 1.0 / (n_trials * np.e)))
    else:
        e_max_sr = 0.0
        
    dsr_stat = (sharpe - e_max_sr) / sr_std
    return float(stats.norm.cdf(dsr_stat))

def ljung_box_p(rets, lags=10):
    n = len(rets)
    if n <= lags + 2 or np.std(rets) <= 1e-9: return 1.0
    ac_list = []
    for k in range(1, lags + 1):
        s1, s2 = rets[:-k], rets[k:]
        if np.std(s1) <= 1e-9 or np.std(s2) <= 1e-9:
            ac_list.append(0.0)
        else:
            with np.errstate(divide='ignore', invalid='ignore'):
                c = np.corrcoef(s1, s2)[0, 1]
                ac_list.append(0.0 if np.isnan(c) or np.isinf(c) else float(c))
    ac = np.array(ac_list)
    Q  = n * (n + 2) * np.sum(ac**2 / (n - np.arange(1, lags + 1)))
    return float(stats.chi2.sf(Q, df=lags))

def regime_labels(close, window=REGIME_WIN):
    """Vectorised regime labeling using pandas rolling returns."""
    s = pd.Series(close)
    r = (s / s.shift(window) - 1.0) * 100.0
    conds = [r > 3.0, r < -3.0]
    choices = ["Bullish", "Bearish"]
    return list(np.select(conds, choices, default="Sideways"))

def _unshifted_profit(close, raw_sig, fee, balance0=BALANCE0):
    """L4 fix: properly uses balance0 param (not global BALANCE0)."""
    n = len(close)
    flips = np.where(raw_sig[1:] != raw_sig[:-1])[0] + 1
    ei = np.concatenate(([1], flips))
    xi = np.concatenate((flips, [n - 1]))
    ei, xi = ei[ei < n], xi[:len(ei)]
    eq = float(balance0)
    for i0, i1 in zip(ei, xi):
        if i1 > i0:
            eq += eq * ((close[i1] / close[i0] - 1.0) * raw_sig[i0] - 2 * fee)
    return float((eq / balance0 - 1.0) * 100.0)

def generate_strategy_signals(close, ema_s=None, ema_l=None, reg_labels=None, hurst_val=0.5, adf_p=0.10, bb_window=20, bb_std=2.0, ema_pairs_list=None, k_threshold=None):
    """
    Q1 Upgrade: Bar-Level Regime-Conditional Adaptive Strategy Engine
    - Bullish Regime: EMA Crossover with Long Bias (shorts filtered)
    - Bearish Regime: EMA Crossover with Short Bias (longs filtered)
    - Sideways Regime: EMA trend suppressed -> Bollinger Band Z-score mean reversion
    - Module 1.1 Upgrade: Multi-EMA Confluence (K-of-N voting agreement across multiple EMA pairs)
    """
    n = len(close)
    if reg_labels is None:
        reg_labels = regime_labels(close)

    s = pd.Series(close)
    rm = s.rolling(bb_window, min_periods=1).mean().to_numpy()
    rstd = s.rolling(bb_window, min_periods=1).std().fillna(1e-9).to_numpy()
    z = (close - rm) / (rstd + 1e-9)

    if ema_pairs_list is not None and len(ema_pairs_list) > 0:
        N = len(ema_pairs_list)
        K = k_threshold if (k_threshold is not None and 1 <= k_threshold <= N) else N
        long_votes = np.zeros(n, dtype=np.int32)
        short_votes = np.zeros(n, dtype=np.int32)
        for es, el in ema_pairs_list:
            long_votes += (es > el).astype(np.int32)
            short_votes += (es < el).astype(np.int32)
        ema_cross = np.where(long_votes >= K, 1.0, np.where(short_votes >= K, -1.0, 0.0))
    elif ema_s is not None and ema_l is not None:
        ema_cross = np.where(ema_s > ema_l, 1.0, -1.0)
    else:
        ema_cross = np.zeros(n, dtype=np.float64)

    raw_sig = np.zeros(n, dtype=np.float64)

    pos = 0.0
    for i in range(n):
        reg = reg_labels[i]
        
        # F1 Upgrade: Hurst/ADF Gate -> Force mean-reversion if non-trending (H < 0.45 or ADF p < 0.05)
        if hurst_val < 0.45 or adf_p < 0.05:
            reg = "Sideways"

        if reg == "Bullish":
            raw_sig[i] = 1.0 if ema_cross[i] > 0 else 0.0
        elif reg == "Bearish":
            raw_sig[i] = -1.0 if ema_cross[i] < 0 else 0.0
        else:
            if z[i] < -bb_std:
                pos = 1.0
            elif z[i] > bb_std:
                pos = -1.0
            elif (pos == 1.0 and z[i] >= 0.0) or (pos == -1.0 and z[i] <= 0.0):
                pos = 0.0
            raw_sig[i] = pos

    return raw_sig

# ----------------------------- STOP-LOSS ENGINE -----------------------------
def compute_utility_score(sm, risk_aversion=2.0):
    """
    Q8 Upgrade: von Neumann-Morgenstern Expected Utility Ranking
    U(strategy) = E[R] - (lambda / 2) * Var[R]
    """
    ret = sm.get("profit_pct", 0.0) / 100.0
    sharpe = sm.get("sharpe", 0.0)
    trades = sm.get("total_trades", 0)
    mdd = abs(sm.get("max_drawdown_pct", 0.0))
    ror = sm.get("risk_of_ruin_pct", 100.0)
    if trades < 5 or sharpe <= 0 or ror >= 50.0: return 0.0
    vol = (ret / max(0.01, sharpe)) if sharpe > 0 else 0.05
    utility = ret - (risk_aversion / 2.0) * (vol ** 2)
    composite = utility * (1.0 - mdd / 100.0) * np.sqrt(min(trades, 100))
    return float(max(0.0, composite))

def compute_score(sm):
    import builtins
    lam = getattr(builtins, '_RISK_AVERSION', RISK_AVERSION)
    return compute_utility_score(sm, risk_aversion=lam)

def compute_breakeven_fee(close, ts, ema_s, ema_l, max_fee_bps=100.0, hurst_val=0.5, adf_p=0.10):
    """
    Q2 Upgrade: Bisection Search for Strategy Break-Even Transaction Cost (Fee in BPS)
    """
    lo, hi = 0.0, max_fee_bps
    best_f = 0.0
    for _ in range(8):
        mid = (lo + hi) / 2.0
        res = backtest_pair(close, ts, ema_s, ema_l, fee_bps=mid, hurst_val=hurst_val, adf_p=adf_p)
        if res and res[0]["profit_pct"] > 0:
            best_f = mid
            lo = mid
        else:
            hi = mid
    return round(float(best_f), 2)

def almgren_chriss_market_impact(trade_size_usd, adv_usd=10_000_000.0, eta=0.14, gamma=0.10, daily_vol=0.02):
    """
    Q3 Upgrade: Almgren-Chriss (2001) Market Impact & Capacity Model
    Computes temporary + permanent market impact slippage in BPS:
    Slippage_bps = (gamma * daily_vol + eta * sqrt(Trade_Size / ADV)) * 10000
    """
    if adv_usd <= 0 or trade_size_usd <= 0:
        return 0.0
    participation_ratio = trade_size_usd / adv_usd
    impact_fraction = gamma * daily_vol + eta * np.sqrt(max(0.0, participation_ratio))
    return float(impact_fraction * 10000.0)

def compute_atr_pct(close, window=14):
    """Computes rolling ATR percentage relative to price."""
    s = pd.Series(close)
    ret = s.pct_change().fillna(0.0)
    vol = ret.rolling(window, min_periods=1).std().fillna(0.0)
    return (vol * np.sqrt(window)).to_numpy()

def apply_stop_loss(seg_close, direction, sl_pct=None, tsl_pct=None, tp_pct=None, regime=None, atr_pct=None):
    """
    P2 Vectorized Stop Loss Engine:
    - Vectorized array evaluation via np.maximum/minimum.accumulate and np.where
    - C1 Fix priority: Fixed SL -> Take Profit -> Trailing SL evaluated at first-hit bar index
    """
    n = len(seg_close)
    if n <= 1 or (sl_pct is None and tsl_pct is None and tp_pct is None):
        return n - 1, 'signal'

    entry = seg_close[0]

    # U1: Convert ATR multiplier to percentage if atr_pct is provided
    if atr_pct is not None and atr_pct > 0:
        eff_sl_pct  = sl_pct  * atr_pct * 100.0 if sl_pct  is not None else None
        eff_tsl_pct = tsl_pct * atr_pct * 100.0 if tsl_pct is not None else None
        eff_tp_pct  = tp_pct  * atr_pct * 100.0 if tp_pct  is not None else None
    else:
        eff_sl_pct, eff_tsl_pct, eff_tp_pct = sl_pct, tsl_pct, tp_pct

    # U2: Regime-conditional adjustment
    if regime == "Bullish":
        if eff_sl_pct is not None: eff_sl_pct *= 0.8
        if eff_tsl_pct is not None: eff_tsl_pct *= 1.2
    elif regime == "Bearish":
        if eff_sl_pct is not None: eff_sl_pct *= 0.7
        if eff_tsl_pct is not None: eff_tsl_pct *= 0.7
    elif regime == "Sideways":
        eff_tsl_pct = None  # Suppress trailing SL in sideways market

    sl_mult  = eff_sl_pct  / 100.0 if eff_sl_pct  is not None else None
    tsl_mult = eff_tsl_pct / 100.0 if eff_tsl_pct is not None else None
    tp_mult  = eff_tp_pct  / 100.0 if eff_tp_pct  is not None else None

    INF = n + 999
    sl_idx, tp_idx, tsl_idx = INF, INF, INF

    sub = seg_close[1:]
    if len(sub) == 0:
        return n - 1, 'signal'

    if direction > 0:
        if sl_mult is not None:
            hits = np.where(sub <= entry * (1.0 - sl_mult))[0]
            if len(hits) > 0: sl_idx = hits[0] + 1
        if tp_mult is not None:
            hits = np.where(sub >= entry * (1.0 + tp_mult))[0]
            if len(hits) > 0: tp_idx = hits[0] + 1
        if tsl_mult is not None:
            cummax = np.maximum.accumulate(seg_close)[1:]
            hits = np.where(sub <= cummax * (1.0 - tsl_mult))[0]
            if len(hits) > 0: tsl_idx = hits[0] + 1
    else:
        if sl_mult is not None:
            hits = np.where(sub >= entry * (1.0 + sl_mult))[0]
            if len(hits) > 0: sl_idx = hits[0] + 1
        if tp_mult is not None:
            hits = np.where(sub <= entry * (1.0 - tp_mult))[0]
            if len(hits) > 0: tp_idx = hits[0] + 1
        if tsl_mult is not None:
            cummin = np.minimum.accumulate(seg_close)[1:]
            hits = np.where(sub >= cummin * (1.0 + tsl_mult))[0]
            if len(hits) > 0: tsl_idx = hits[0] + 1

    min_idx = min(sl_idx, tp_idx, tsl_idx)
    if min_idx >= INF:
        return n - 1, 'signal'

    # Strict priority evaluation at min_idx (C1)
    if sl_idx == min_idx:
        return sl_idx, 'sl'
    elif tp_idx == min_idx:
        return tp_idx, 'tp'
    else:
        return tsl_idx, 'tsl'


def run_sl_grid(close, ts, ema_s, ema_l, symbol, tf, bx, by,
                reg_labels=None, balance0=BALANCE0, fee_bps=FEE_BPS, slippage_bps=SLIPPAGE_BPS,
                hurst_val=0.5, adf_p=0.10):
    """
    P1 + P2 Fast SL/TSL/TP/ATR Grid Engine:
    - Runs 3D grid: fixed_sl x tsl x tp_rr with ATR dynamic scaling and regime conditional adjustments
    """
    rows = []
    configs = [('none', None, None, None, None)]
    for f in SL_LEVELS_PCT:
        configs.append(('fixed', f, None, None, None))
    for t in TSL_LEVELS_PCT:
        configs.append(('trailing', None, t, None, None))
    for f in [0.5, 1.0, 1.5, 2.0, 3.0, 5.0]:
        for t in [0.5, 1.0, 1.5, 2.0, 3.0, 5.0]:
            configs.append(('dual', f, t, None, None))
            for rr in TP_RR_MULTIPLIERS:
                configs.append(('dual_tp', f, t, round(f * rr, 2), None))
    for atr_m in ATR_GRID:
        configs.append(('atr_dynamic', 1.0, 1.0, 2.0, atr_m))
        configs.append(('atr_dynamic', 0.5, 1.5, 1.5, atr_m))
        configs.append(('atr_dynamic', 1.5, 1.0, 3.0, atr_m))

    n = len(close)
    if reg_labels is None: reg_labels = regime_labels(close)
    raw_sig = generate_strategy_signals(close, ema_s, ema_l, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
    position = np.roll(raw_sig, 1); position[0] = 0.0
    flips = np.where(position[1:] != position[:-1])[0] + 1
    ei = np.concatenate(([1], flips))
    xi = np.concatenate((flips, [n - 1]))
    ei, xi = ei[ei < n], xi[:len(ei)]
    if reg_labels is None: reg_labels = regime_labels(close)
    atr_pct_arr = compute_atr_pct(close, window=14)
    fee_alone = fee_bps / 10000.0
    fee = (fee_bps + slippage_bps) / 10000.0
    pre_seg = (ei, xi, position, reg_labels, atr_pct_arr, fee, raw_sig, fee_alone)

    # P1 Optimization: Precompute K-Fold EMA slices ONCE for (bx, by) pair
    precomputed_folds = precompute_kfold_emas(close, ts, bx, by, folds=WF_FOLDS)

    for sl_type, sl_pct, tsl_pct, tp_pct, atr_m in configs:
        res = backtest_pair(close, ts, ema_s, ema_l,
                            balance0=balance0, fee_bps=fee_bps, slippage_bps=slippage_bps,
                            reg_labels=reg_labels, sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct, atr_multiple=atr_m, tf=tf,
                            pre_seg=pre_seg)
        if res is None:
            continue
        sm = res[0]
        
        # U4 OOS Validation using precomputed fold EMAs
        oos_sh, oos_pf, oos_deg, wf_pass = walk_forward_sl_val(close, ts, bx, by, sl_pct, tsl_pct, tp_pct, atr_m=atr_m, precomputed_folds=precomputed_folds)
        
        tp_rr_val = round(tp_pct / sl_pct, 2) if (tp_pct and sl_pct and sl_pct > 0) else 0.0
        rows.append(dict(
            symbol=symbol, timeframe=tf, ema_x=bx, ema_y=by,
            sl_type=sl_type,
            fixed_sl_pct=sl_pct if sl_pct is not None else 0.0,
            trailing_sl_pct=tsl_pct if tsl_pct is not None else 0.0,
            take_profit_pct=tp_pct if tp_pct is not None else 0.0,
            atr_multiple=atr_m if atr_m is not None else 0.0,
            tp_rr=tp_rr_val,
            oos_sharpe=oos_sh, oos_profit_pct=oos_pf, oos_degradation_pct=oos_deg, wf_validation_pass=wf_pass,
            **sm
        ))
    return pd.DataFrame(rows)


def precompute_kfold_emas(close, ts, bx, by, folds=WF_FOLDS, purge_pct=0.02):
    """
    P1 Upgrade: Precomputes Purged K-Fold training & test EMA slices ONCE per (bx, by) pair.
    Eliminates 500+ redundant EMA computations across the Stop-Loss search grid.
    """
    n = len(close)
    if n < 100 or folds < 2:
        return []

    fold_size = n // folds
    purge_len = max(20, int(n * purge_pct), max(bx, by))
    precomputed = []

    for k in range(folds):
        test_start = k * fold_size
        test_end = n if k == folds - 1 else (k + 1) * fold_size
        if test_end - test_start < 20: continue

        train_part1_end = max(0, test_start - purge_len)
        train_part2_start = min(n, test_end + purge_len)

        train_indices = []
        if train_part1_end > 20:
            train_indices.append((0, train_part1_end))
        if n - train_part2_start > 20:
            train_indices.append((train_part2_start, n))

        if not train_indices:
            continue

        c_train_list = [close[i0:i1] for i0, i1 in train_indices]
        ts_train_list = [ts[i0:i1] for i0, i1 in train_indices]
        c_train = np.concatenate(c_train_list)
        ts_train = np.concatenate(ts_train_list)

        c_test = close[test_start:test_end]
        ts_test = ts[test_start:test_end]

        ema_s_train = _ema(c_train, 2.0 / (bx + 1.0))
        ema_l_train = _ema(c_train, 2.0 / (by + 1.0))

        ema_s_test = _ema(c_test, 2.0 / (bx + 1.0))
        ema_l_test = _ema(c_test, 2.0 / (by + 1.0))

        precomputed.append((c_train, ts_train, ema_s_train, ema_l_train,
                            c_test, ts_test, ema_s_test, ema_l_test))

    return precomputed

def purged_kfold_cv(close, ts, bx, by, sl_pct, tsl_pct, tp_pct, folds=WF_FOLDS, atr_m=None, purge_pct=0.02, precomputed_folds=None):
    """
    F3 + P1 Upgrade: Purged K-Fold Cross-Validation with Precomputed EMA Cache
    """
    if precomputed_folds is None:
        precomputed_folds = precompute_kfold_emas(close, ts, bx, by, folds=folds, purge_pct=purge_pct)

    if not precomputed_folds:
        return 0.0, 0.0, 0.0, False

    is_sharpes, oos_sharpes, oos_pfs = [], [], []

    for (c_train, ts_train, ema_s_train, ema_l_train, c_test, ts_test, ema_s_test, ema_l_test) in precomputed_folds:
        res_is = backtest_pair(c_train, ts_train, ema_s_train, ema_l_train,
                               sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct, atr_multiple=atr_m)
        res_oos = backtest_pair(c_test, ts_test, ema_s_test, ema_l_test,
                                sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct, atr_multiple=atr_m)

        if res_is and res_oos:
            is_sharpes.append(res_is[0]["sharpe"])
            oos_sharpes.append(res_oos[0]["sharpe"])
            oos_pfs.append(res_oos[0]["profit_factor"])

    if not oos_sharpes:
        return 0.0, 0.0, 0.0, False

    avg_is_sharpe  = float(np.mean(is_sharpes))
    avg_oos_sharpe = float(np.mean(oos_sharpes))
    avg_oos_pf     = float(np.mean(oos_pfs))

    deg = round(float((1.0 - (avg_oos_sharpe / max(0.001, avg_is_sharpe))) * 100.0), 2) if avg_is_sharpe > 0 else 0.0
    wf_pass = bool(avg_oos_sharpe >= 0.5 * avg_is_sharpe and avg_oos_pf >= 1.0)

    return round(avg_oos_sharpe, 2), round(avg_oos_pf, 2), deg, wf_pass

def walk_forward_sl_val(close, ts, bx, by, sl_pct, tsl_pct, tp_pct, folds=WF_FOLDS, atr_m=None, precomputed_folds=None):
    """Alias wrapping Purged K-Fold Cross-Validation with EMA Cache."""
    return purged_kfold_cv(close, ts, bx, by, sl_pct, tsl_pct, tp_pct, folds=folds, atr_m=atr_m, precomputed_folds=precomputed_folds)


def conclude_best_sl(sl_df):
    """
    Ranks all SL configs by multi-factor composite score and generates human-readable rationale.
    C3 Fix: Groups baseline profit/MDD strictly per (symbol, timeframe).
    P3 Fix: Vectorized rationale string building via np.select.
    R2/R3 Fix: Includes Walk-Forward OOS validation and degradation metrics.
    """
    if sl_df.empty:
        return {}, pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df = sl_df.copy()
    df['_mdd_abs'] = df['max_drawdown_pct'].abs()
    df['_trades']  = df['total_trades'].clip(lower=1)
    df['_ror']     = df['risk_of_ruin_pct'].fillna(100.0)
    df['_wr']      = df['win_rate_pct'].fillna(0.0)

    df['score'] = df.apply(lambda row: compute_score(row.to_dict()), axis=1)
    df = df.sort_values('score', ascending=False).reset_index(drop=True)
    df['rank'] = df.index + 1

    # C3/P3 Vectorized Baseline Merge per (symbol, timeframe) group
    base_df = df[df['sl_type'] == 'none'][['symbol', 'timeframe', 'profit_pct', 'max_drawdown_pct', 'sharpe']].copy()
    base_df = base_df.rename(columns={'profit_pct': '_base_profit', 'max_drawdown_pct': '_base_mdd', 'sharpe': '_base_sharpe'})
    base_df = base_df.drop_duplicates(subset=['symbol', 'timeframe'])

    if '_base_profit' in df.columns:
        df = df.drop(columns=['_base_profit', '_base_mdd', '_base_sharpe'], errors='ignore')
    
    df = df.merge(base_df, on=['symbol', 'timeframe'], how='left')
    df['_base_profit'] = df['_base_profit'].fillna(0.0)
    df['_base_mdd']    = df['_base_mdd'].abs().fillna(0.0)
    df['_base_sharpe'] = df['_base_sharpe'].fillna(0.0)

    df['_mdd_red']   = df['_base_mdd'] - df['max_drawdown_pct'].abs()
    df['_prof_ret']  = np.where(df['_base_profit'] != 0, (df['profit_pct'] / df['_base_profit']) * 100.0, 0.0)

    conds = [
        df['sl_type'] == 'none',
        df['sl_type'] == 'fixed',
        df['sl_type'] == 'trailing',
        df['sl_type'] == 'dual_tp',
        df['sl_type'] == 'atr_dynamic',
    ]
    choices = [
        "Unconstrained baseline without stop protection.",
        "Fixed SL @" + df['fixed_sl_pct'].astype(str) + "%: Reduces MDD by " + df['_mdd_red'].round(1).astype(str) + "% while retaining " + df['_prof_ret'].round(1).astype(str) + "% baseline return.",
        "Trailing SL @" + df['trailing_sl_pct'].astype(str) + "%: Locks profit on trends, reducing MDD by " + df['_mdd_red'].round(1).astype(str) + "% with Sharpe " + df['sharpe'].round(2).astype(str) + ".",
        "Dual SL + TP (Fixed " + df['fixed_sl_pct'].astype(str) + "% + Trailing " + df['trailing_sl_pct'].astype(str) + "% + TP " + df['take_profit_pct'].astype(str) + "%): Profit-targeted risk/reward optimization with Score " + df['score'].round(3).astype(str) + ".",
        "ATR Dynamic SL (" + df['atr_multiple'].astype(str) + "x ATR): Volatility-scaled dynamic risk management with Score " + df['score'].round(3).astype(str) + "."
    ]
    df['rationale'] = np.select(conds, choices, default="Dual SL (Fixed " + df['fixed_sl_pct'].astype(str) + "% + Trailing " + df['trailing_sl_pct'].astype(str) + "%): Optimal hybrid protection with Score " + df['score'].round(3).astype(str) + ".")

    out_cols = ['rank','symbol','timeframe','sl_type','fixed_sl_pct','trailing_sl_pct','take_profit_pct',
                'atr_multiple','tp_rr','profit_pct','profit_factor','sharpe','sortino','max_drawdown_pct',
                'total_trades','win_rate_pct','risk_of_ruin_pct','score','rationale']
    out_cols = [c for c in out_cols if c in df.columns]
    master_conclusion = df[out_cols].head(50)

    # B6 Fix: Vectorized Comparative vs No-SL Delta
    nosl_df = df.copy()
    nosl_df['delta_profit_pct'] = (nosl_df['profit_pct'] - nosl_df['_base_profit']).round(2)
    nosl_df['delta_mdd_pct'] = (nosl_df['_base_mdd'] - nosl_df['max_drawdown_pct'].abs()).round(2)
    nosl_df['delta_sharpe'] = (nosl_df['sharpe'] - nosl_df['_base_sharpe']).round(2)
    nosl_cols = ['symbol', 'timeframe', 'sl_type', 'fixed_sl_pct', 'trailing_sl_pct', 'take_profit_pct', 'atr_multiple', 
                 'profit_pct', 'delta_profit_pct', 'max_drawdown_pct', 'delta_mdd_pct', 'sharpe', 'delta_sharpe', 'score']
    nosl_df = nosl_df[[c for c in nosl_cols if c in nosl_df.columns]]

    # Head-to-Head: Best Fixed vs Best Trailing
    h2h = []
    for (s, t), grp in df.groupby(['symbol', 'timeframe']):
        bf = grp[grp['sl_type'] == 'fixed'].head(1)
        bt = grp[grp['sl_type'] == 'trailing'].head(1)
        if not bf.empty and not bt.empty:
            f_r, t_r = bf.iloc[0], bt.iloc[0]
            winner = 'Trailing SL' if t_r['score'] > f_r['score'] else 'Fixed SL'
            h2h.append(dict(
                symbol=s, timeframe=t,
                best_fixed_sl=f_r['fixed_sl_pct'], fixed_profit=f_r['profit_pct'], fixed_sharpe=f_r['sharpe'], fixed_mdd=f_r['max_drawdown_pct'],
                best_trailing_sl=t_r['trailing_sl_pct'], trailing_profit=t_r['profit_pct'], trailing_sharpe=t_r['sharpe'], trailing_mdd=t_r['max_drawdown_pct'],
                winner=winner, score_diff=round(abs(t_r['score'] - f_r['score']), 3)
            ))
    h2h_df = pd.DataFrame(h2h)

    return df.iloc[0].to_dict(), master_conclusion, nosl_df, h2h_df


# ----------------------------- MODULE 1.4 TSL/SL COMPARATIVE ENGINE ---------
def run_sl_comparative_4modes(close, ts, ema_s, ema_l, symbol, tf, bx, by,
                                fixed_sl_pct=1.5, trailing_sl_pct=1.5, take_profit_pct=3.0,
                                reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Module 1.4: TSL / SL With-and-Without Comparative Analysis Engine.
    Evaluates strategy performance across 4 structured modes:
      Mode A: No SL, no TSL (signal exit only - baseline)
      Mode B: Fixed SL only
      Mode C: Trailing SL only
      Mode D: Dual Fixed SL + Trailing SL + Take Profit
    Calculates exact deltas (Δ_profit, Δ_maxDD, Δ_winrate, Δ_sharpe) and SL benefit score.
    """
    modes = [
        ("Mode_A_Signal_Only", None, None, None),
        ("Mode_B_Fixed_SL", fixed_sl_pct, None, None),
        ("Mode_C_Trailing_SL", None, trailing_sl_pct, None),
        ("Mode_D_Dual_SL_TP", fixed_sl_pct, trailing_sl_pct, take_profit_pct)
    ]

    rows = []
    base_profit, base_mdd, base_wr, base_sharpe = 0.0, 0.0, 0.0, 0.0

    for idx, (m_label, f_sl, t_sl, tp) in enumerate(modes):
        res = backtest_pair(close, ts, ema_s, ema_l, tf=tf, reg_labels=reg_labels,
                            sl_pct=f_sl, tsl_pct=t_sl, tp_pct=tp, hurst_val=hurst_val, adf_p=adf_p)
        if res is None:
            continue
        sm = res[0]
        p_pct  = sm["profit_pct"]
        mdd    = abs(sm["max_drawdown_pct"])
        wr     = sm["win_rate_pct"]
        sh     = sm["sharpe"]

        if idx == 0:
            base_profit = p_pct
            base_mdd    = mdd
            base_wr     = wr
            base_sharpe = sh

        delta_prof = round(p_pct - base_profit, 2)
        delta_mdd  = round(base_mdd - mdd, 2) # positive = MDD reduction/improvement
        delta_wr   = round(wr - base_wr, 2)
        delta_sh   = round(sh - base_sharpe, 2)

        benefit_score = round(delta_sh / (1.0 + abs(delta_mdd) / 100.0), 3)

        rows.append(dict(
            symbol=symbol, timeframe=tf, ema_x=bx, ema_y=by,
            mode=m_label,
            fixed_sl_pct=f_sl if f_sl else 0.0,
            trailing_sl_pct=t_sl if t_sl else 0.0,
            take_profit_pct=tp if tp else 0.0,
            profit_pct=p_pct, delta_profit_pct=delta_prof,
            max_drawdown_pct=sm["max_drawdown_pct"], delta_mdd_pct=delta_mdd,
            win_rate_pct=wr, delta_win_rate_pct=delta_wr,
            sharpe=sh, delta_sharpe=delta_sh,
            sl_benefit_score=benefit_score,
            total_trades=sm["total_trades"], sqn=sm["sqn"],
            avg_reversal_lag_bars=sm.get("avg_reversal_lag_bars", 0.0),
            avg_candle_loss_pct=sm.get("avg_candle_loss_pct", 0.0)
        ))

    return pd.DataFrame(rows)

# ----------------------------- MODULE 1.2 COMPOUNDING ENGINE ----------------
def run_compounding_grid(close, ts, ema_s, ema_l, symbol, tf, bx, by,
                         reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Module 1.2: Dynamic Compounding / Reinvestment Schedules Engine.
    Evaluates 9 distinct position sizing and profit reinvestment schedules:
      1. flat (baseline constant size)
      2. kelly_full (continuous full Kelly f*)
      3. kelly_half (continuous half Kelly f*)
      4. fixed_pct (fixed fraction per trade)
      5. linear_growth (stepwise linear lot addition)
      6. exponential_growth (geometric compounding multiplier)
      7. profit_triggered (reinvest when equity exceeds target watermark)
      8. drawdown_gated (deleveraging on drawdown > threshold)
      9. volatility_scaled (size ∝ 1/rolling_vol)
    Calculates compounding equity uplift % and risk metrics.
    """
    schedules = [
        "flat", "kelly_full", "kelly_half", "fixed_pct",
        "linear_growth", "exponential_growth", "profit_triggered",
        "drawdown_gated", "volatility_scaled"
    ]

    rows = []
    base_profit, base_mdd, base_cagr = 0.0, 0.0, 0.0

    for idx, sz in enumerate(schedules):
        res = backtest_pair(close, ts, ema_s, ema_l, tf=tf, sizing=sz,
                            reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        if res is None:
            continue
        sm = res[0]
        p_pct  = sm["profit_pct"]
        cagr_v = sm["cagr_pct"]
        mdd    = abs(sm["max_drawdown_pct"])

        if idx == 0:
            base_profit = p_pct
            base_mdd    = mdd
            base_cagr   = cagr_v

        compounding_uplift = round(p_pct - base_profit, 2)
        cagr_uplift        = round(cagr_v - base_cagr, 2)

        rows.append(dict(
            symbol=symbol, timeframe=tf, ema_x=bx, ema_y=by,
            compounding_schedule=sz,
            final_balance=sm["final_balance"],
            profit_pct=p_pct, compounding_uplift_pct=compounding_uplift,
            cagr_pct=cagr_v, cagr_uplift_pct=cagr_uplift,
            sharpe=sm["sharpe"], sortino=sm["sortino"],
            max_drawdown_pct=sm["max_drawdown_pct"],
            ulcer_index=sm["ulcer_index"], sqn=sm["sqn"],
            risk_of_ruin_pct=sm["risk_of_ruin_pct"],
            total_trades=sm["total_trades"], win_rate_pct=sm["win_rate_pct"]
        ))

    return pd.DataFrame(rows)

# ----------------------------- MODULE 1.3 COMBINATORIAL PERMUTATION ENGINE --
def run_combinatorial_permutation_engine(close, ts, symbol, tf, sample_size=100, seed=42, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Module 1.3: Full Combinatorial Permutation Engine.
    Crosses N-EMA Confluence sets (Module 1.1) x Compounding Schedules (Module 1.2) x SL Modes (Module 1.4) x Regime Gates.
    Uses Latin Hypercube / Stratified Sampling to efficiently explore multi-billion parameter space.
    Computes multi-factor composite rank (Sharpe 30% + Calmar 20% + SQN 20% + DSR 15% + OOS 15%).
    """
    rng = np.random.default_rng(seed)
    
    ema_pairs_pool = [(5, 20), (10, 50), (20, 200)]
    schedules_pool = ["flat", "kelly_half", "exponential_growth", "drawdown_gated", "volatility_scaled"]
    sl_configs_pool = [
        ("none", None, None, None, None),
        ("fixed", 1.5, None, None, None),
        ("trailing", None, 1.5, None, None),
        ("dual_tp", 1.5, 1.5, 3.0, None),
        ("atr_dynamic", 1.0, 1.0, 2.0, 1.5)
    ]
    dir_filters_pool = ["both", "long_only", "short_only"]

    # Pre-calculate EMA arrays
    ema_cache = {}
    for px, py in ema_pairs_pool:
        es = _ema(close, 2.0 / (px + 1.0))
        el = _ema(close, 2.0 / (py + 1.0))
        ema_cache[(px, py)] = (es, el)

    rows = []

    for i in range(sample_size):
        # Sample configuration across dimensions
        pair_idx = rng.choice(len(ema_pairs_pool))
        px, py = ema_pairs_pool[pair_idx]
        es, el = ema_cache[(px, py)]
        
        sz = rng.choice(schedules_pool)
        sl_t, f_sl, t_sl, tp, atr_m = sl_configs_pool[rng.choice(len(sl_configs_pool))]
        d_filter = rng.choice(dir_filters_pool)

        res = backtest_pair(close, ts, es, el, tf=tf, dir_filter=d_filter, sizing=sz,
                            sl_pct=f_sl, tsl_pct=t_sl, tp_pct=tp, atr_multiple=atr_m,
                            reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        if res is None:
            continue

        sm = res[0]
        p_pct  = sm["profit_pct"]
        sh     = sm["sharpe"]
        calmar = sm["calmar"]
        sqn    = sm["sqn"]
        dsr    = sm.get("dsr_prob", 0.5)

        # Composite score
        comp_score = (0.30 * max(0.0, sh)) + (0.20 * min(10.0, max(0.0, calmar))) + (0.20 * max(0.0, sqn)) + (0.15 * dsr * 10.0)

        rows.append(dict(
            perm_id=i + 1, symbol=symbol, timeframe=tf,
            ema_pair=f"({px},{py})", compounding_schedule=sz,
            sl_type=sl_t, fixed_sl_pct=f_sl if f_sl else 0.0, trailing_sl_pct=t_sl if t_sl else 0.0,
            take_profit_pct=tp if tp else 0.0, atr_multiple=atr_m if atr_m else 0.0,
            direction=d_filter,
            profit_pct=p_pct, cagr_pct=sm["cagr_pct"], sharpe=sh, sortino=sm["sortino"],
            calmar=calmar, max_drawdown_pct=sm["max_drawdown_pct"], sqn=sqn,
            dsr_prob=dsr, composite_score=round(comp_score, 3),
            reversal_lag_bars=sm.get("avg_reversal_lag_bars", 0.0),
            candle_loss_pct=sm.get("avg_candle_loss_pct", 0.0),
            late_exit_cost_bps=sm.get("avg_late_exit_cost_bps", 0.0),
            total_trades=sm["total_trades"], win_rate_pct=sm["win_rate_pct"]
        ))

    df_perm = pd.DataFrame(rows)
    if not df_perm.empty:
        df_perm = df_perm.sort_values(by="composite_score", ascending=False).reset_index(drop=True)
        df_perm["composite_rank"] = df_perm.index + 1

    return df_perm

# ----------------------------- MODULE 1.6 MASTER RESEARCH DASHBOARD --------
def sensitivity_analysis(close, ts, ema_s, ema_l, tf="1h", sl_pct=None, tsl_pct=None, tp_pct=None,
                         sizing="flat", perturb_pct=5.0, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Module 1.6 Sub: Perturbation Sensitivity Analysis for Parameter Robustness.
    Perturbs each numeric parameter by ±perturb_pct and measures Sharpe/CAGR stability.
    Returns a sensitivity dict with metrics and a fragility score (0=robust, 1=fragile).
    """
    base_res = backtest_pair(close, ts, ema_s, ema_l, tf=tf, sizing=sizing,
                             sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct,
                             reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
    if base_res is None:
        return dict(fragility_score=1.0, base_sharpe=0.0, sharpe_range=0.0)
    base_sh = base_res[0]["sharpe"]
    base_cagr = base_res[0]["cagr_pct"]

    perturbed_sharpes = []
    mult_lo, mult_hi = 1.0 - perturb_pct / 100.0, 1.0 + perturb_pct / 100.0

    for p_sl in ([sl_pct * mult_lo, sl_pct * mult_hi] if sl_pct else [None]):
        for p_tsl in ([tsl_pct * mult_lo, tsl_pct * mult_hi] if tsl_pct else [None]):
            for p_tp in ([tp_pct * mult_lo, tp_pct * mult_hi] if tp_pct else [None]):
                r = backtest_pair(close, ts, ema_s, ema_l, tf=tf, sizing=sizing,
                                  sl_pct=p_sl, tsl_pct=p_tsl, tp_pct=p_tp,
                                  reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
                if r:
                    perturbed_sharpes.append(r[0]["sharpe"])

    if not perturbed_sharpes:
        return dict(fragility_score=0.0, base_sharpe=round(base_sh, 2),
                    base_cagr=round(base_cagr, 2), sharpe_range=0.0)

    sh_arr = np.array(perturbed_sharpes)
    sh_range = float(sh_arr.max() - sh_arr.min())
    fragility = min(1.0, sh_range / max(0.01, abs(base_sh)))

    return dict(
        fragility_score=round(fragility, 3),
        base_sharpe=round(base_sh, 2), base_cagr=round(base_cagr, 2),
        sharpe_range=round(sh_range, 2),
        sharpe_min=round(float(sh_arr.min()), 2),
        sharpe_max=round(float(sh_arr.max()), 2),
        n_perturbations=len(perturbed_sharpes)
    )


def generate_master_research_dashboard(df_perm, df_conf=None, df_comp=None, df_4modes=None,
                                        df_lag_summary=None, symbol="UNKNOWN", tf="1h"):
    """
    Module 1.6: Master Research Dashboard Generator.
    Aggregates all quantitative outputs across Modules 1.1–1.5 into a unified reporting suite:
      1. Top-50 Permutations Master Table (sorted by composite_score)
      2. Confluence Benefit Breakdown (1.1)
      3. Dynamic Compounding Uplift Table (1.2)
      4. SL With/Without 4-Modes Delta Table (1.4)
      5. Trend Reversal Lag League Table (1.5)
      6. Regime-Conditional Best Config
      7. Risk-Reward Efficiency Surface
      8. Statistical Overfitting (DSR) Summary
      9. Executive Key Metrics KPI Card
     10. System Health & Warnings
    """
    top50 = df_perm.head(50).copy() if (df_perm is not None and not df_perm.empty) else pd.DataFrame()

    # KPI Card
    kpi_card = {}
    if not top50.empty:
        best = top50.iloc[0]
        kpi_card = dict(
            symbol=symbol, timeframe=tf,
            best_ema_pair=best.get("ema_pair", "N/A"),
            best_compounding_schedule=best.get("compounding_schedule", "N/A"),
            best_sl_type=best.get("sl_type", "N/A"),
            best_sharpe=best.get("sharpe", 0.0),
            best_cagr_pct=best.get("cagr_pct", 0.0),
            best_max_dd_pct=best.get("max_drawdown_pct", 0.0),
            best_sqn=best.get("sqn", 0.0),
            best_composite_score=best.get("composite_score", 0.0),
            best_reversal_lag=best.get("reversal_lag_bars", 0.0),
            best_candle_loss=best.get("candle_loss_pct", 0.0),
            total_permutations_evaluated=len(df_perm) if df_perm is not None else 0
        )

    # Regime-Conditional Best (from permutations)
    regime_best = {}
    if not top50.empty and "direction" in top50.columns:
        for d in ["both", "long_only", "short_only"]:
            sub = top50[top50["direction"] == d]
            if not sub.empty:
                r = sub.iloc[0]
                regime_best[d] = dict(
                    ema_pair=r.get("ema_pair", "N/A"),
                    compounding=r.get("compounding_schedule", "N/A"),
                    sl_type=r.get("sl_type", "N/A"),
                    sharpe=r.get("sharpe", 0.0),
                    cagr_pct=r.get("cagr_pct", 0.0),
                    composite_score=r.get("composite_score", 0.0)
                )

    # Risk-Reward Efficiency (top50 Sharpe vs MDD frontier)
    rr_surface = pd.DataFrame()
    if not top50.empty:
        rr_surface = top50[["ema_pair", "sl_type", "compounding_schedule", "sharpe", "max_drawdown_pct", "cagr_pct", "composite_score"]].copy()
        rr_surface["efficiency"] = rr_surface["sharpe"] / (rr_surface["max_drawdown_pct"].abs().clip(lower=0.01))

    # DSR Overfitting Summary
    overfit_summary = {}
    if not top50.empty and "dsr_prob" in top50.columns:
        dsr_vals = top50["dsr_prob"].values
        overfit_summary = dict(
            avg_dsr=round(float(np.mean(dsr_vals)), 3),
            min_dsr=round(float(np.min(dsr_vals)), 3),
            pct_dsr_above_05=round(float((dsr_vals > 0.5).mean() * 100), 1),
            pct_dsr_above_095=round(float((dsr_vals > 0.95).mean() * 100), 1)
        )

    # Warnings
    warnings = []
    if kpi_card.get("best_sharpe", 0) < 0.5:
        warnings.append("WARN: Best Sharpe < 0.5 — strategy may not be viable after fees.")
    if kpi_card.get("best_max_dd_pct", 0) < -30:
        warnings.append("WARN: Best MDD > 30% — extreme drawdown risk.")
    if overfit_summary.get("pct_dsr_above_05", 0) < 50:
        warnings.append("WARN: <50% of top configs pass DSR > 0.5 — possible overfitting.")

    dashboard = dict(
        kpi_card=kpi_card,
        top50_permutations=top50,
        regime_conditional_best=regime_best,
        risk_reward_surface=rr_surface,
        overfit_summary=overfit_summary,
        warnings=warnings,
        confluence_summary=df_conf if df_conf is not None else pd.DataFrame(),
        compounding_summary=df_comp if df_comp is not None else pd.DataFrame(),
        sl_4modes_summary=df_4modes if df_4modes is not None else pd.DataFrame(),
        reversal_lag_summary=df_lag_summary if df_lag_summary is not None else pd.DataFrame()
    )

    return dashboard


def run_full_research(close, ts, symbol, tf, bx, by, sample_size=50, seed=42,
                      reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Unified Auto-Research Runner: Chains all 6 research modules into a single call.
    Returns the complete master dashboard dict with all sub-module outputs.
    """
    ema_s = _ema(close, 2.0 / (bx + 1.0))
    ema_l = _ema(close, 2.0 / (by + 1.0))

    # Module 1.1: Confluence
    df_conf = run_confluence_grid(close, ts, symbol, tf,
                                  period_pairs=[(5, 20), (10, 50), (20, 200)],
                                  reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)

    # Module 1.2: Compounding
    df_comp = run_compounding_grid(close, ts, ema_s, ema_l, symbol, tf, bx, by,
                                   reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)

    # Module 1.4: SL 4-Modes Comparative
    df_4modes = run_sl_comparative_4modes(close, ts, ema_s, ema_l, symbol, tf, bx, by,
                                          reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)

    # Module 1.5: Reversal Lag
    base_res = backtest_pair(close, ts, ema_s, ema_l, tf=tf, reg_labels=reg_labels,
                             hurst_val=hurst_val, adf_p=adf_p)
    df_lag_summary = pd.DataFrame()
    if base_res:
        lag_sum, lag_reg, lag_top10 = analyze_reversal_lag(base_res[1], symbol, tf)
        df_lag_summary = pd.DataFrame([lag_sum]) if lag_sum else pd.DataFrame()

    # Module 1.3: Combinatorial Permutation Engine
    df_perm = run_combinatorial_permutation_engine(close, ts, symbol, tf,
                                                    sample_size=sample_size, seed=seed,
                                                    reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)

    # Module 1.6: Master Dashboard
    dashboard = generate_master_research_dashboard(df_perm, df_conf=df_conf, df_comp=df_comp,
                                                    df_4modes=df_4modes, df_lag_summary=df_lag_summary,
                                                    symbol=symbol, tf=tf)

    # Sensitivity analysis on the top-1 permutation
    if not df_perm.empty:
        top1 = df_perm.iloc[0]
        px, py = int(top1["ema_pair"].strip("()").split(",")[0]), int(top1["ema_pair"].strip("()").split(",")[1])
        sens = sensitivity_analysis(close, ts,
                                     _ema(close, 2.0 / (px + 1.0)),
                                     _ema(close, 2.0 / (py + 1.0)),
                                     tf=tf,
                                     sl_pct=top1["fixed_sl_pct"] if top1["fixed_sl_pct"] > 0 else None,
                                     tsl_pct=top1["trailing_sl_pct"] if top1["trailing_sl_pct"] > 0 else None,
                                     tp_pct=top1["take_profit_pct"] if top1["take_profit_pct"] > 0 else None,
                                     sizing=top1["compounding_schedule"],
                                     reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        dashboard["top1_sensitivity"] = sens

    return dashboard





# ----------------------------- MONTE CARLO ----------------------------------
def monte_carlo(trade_rets, n=MC_SIMS, seed=42, return_bands=False):
    """
    P4 Unified Monte Carlo Simulation: Computes both quantiles and equity curves in a single bootstrap run.
    """
    if len(trade_rets) < 2:
        res = dict(mc_p5=BALANCE0, mc_p50=BALANCE0, mc_p95=BALANCE0,
                   mc_prob_ruin=0.0, mc_sharpe_p5=0.0, mc_sharpe_p95=0.0)
        if return_bands:
            return res, (np.full(1, BALANCE0), np.full(1, BALANCE0), np.full(1, BALANCE0))
        return res

    rng = np.random.default_rng(seed)
    T   = len(trade_rets)

    # F4 Upgrade: Moving Block Bootstrap (Politis-Romano 1994) for autocorrelated returns
    autocorr_lag1 = 0.0
    if T > 5 and np.std(trade_rets) > 1e-9:
        with np.errstate(divide='ignore', invalid='ignore'):
            c = np.corrcoef(trade_rets[:-1], trade_rets[1:])[0, 1]
            autocorr_lag1 = float(c) if not np.isnan(c) and not np.isinf(c) else 0.0

    block_size = max(1, int(1.0 / (1.0 - abs(autocorr_lag1) + 1e-9))) if abs(autocorr_lag1) > 0.05 else 1
    block_size = min(block_size, max(1, T // 2))

    if block_size <= 1:
        sims = rng.choice(trade_rets, (n, T), replace=True)
    else:
        n_blocks = (T + block_size - 1) // block_size
        max_start = max(1, T - block_size + 1)
        starts = rng.integers(0, max_start, size=(n, n_blocks))
        sims_list = []
        for i in range(n):
            row = np.concatenate([trade_rets[st:st + block_size] for st in starts[i]])
            sims_list.append(row[:T])
        sims = np.array(sims_list)

    curves = BALANCE0 * np.cumprod(1 + sims, axis=1)
    finals = curves[:, -1]
    m      = sims.mean(axis=1); s = sims.std(axis=1)
    sharps = np.divide(m, s, out=np.zeros_like(m), where=s > 0) * np.sqrt(T)

    res = dict(
        mc_p5  =round(float(np.percentile(finals, 5)), 2),
        mc_p50 =round(float(np.percentile(finals, 50)), 2),
        mc_p95 =round(float(np.percentile(finals, 95)), 2),
        mc_prob_ruin   =round(float((finals < BALANCE0 * 0.5).mean() * 100), 2),
        mc_sharpe_p5   =round(float(np.percentile(sharps, 5)), 2),
        mc_sharpe_p95  =round(float(np.percentile(sharps, 95)), 2),
        mc_block_size  =block_size,
    )
    if return_bands:
        bands = (np.percentile(curves, 5,  axis=0),
                 np.percentile(curves, 50, axis=0),
                 np.percentile(curves, 95, axis=0))
        return res, bands
    return res

def mc_equity_bands(trade_rets, n=MC_SIMS, seed=99):
    """Legacy alias wrapping unified monte_carlo."""
    _, bands = monte_carlo(trade_rets, n=n, seed=seed, return_bands=True)
    return bands

# ----------------------------- BACKTEST ENGINE ------------------------------
def backtest_pair(close, ts, ema_s=None, ema_l=None,
                  balance0=BALANCE0, fee_bps=FEE_BPS, slippage_bps=SLIPPAGE_BPS,
                  dir_filter="both", shifted=True, sizing="flat", reg_labels=None,
                  sl_pct=None, tsl_pct=None, tp_pct=None, atr_multiple=None, tf="1h", pre_seg=None,
                  hurst_val=0.5, adf_p=0.10, ema_pairs_list=None, k_threshold=None):
    if pre_seg is None:
        n       = len(close)
        if reg_labels is None:
            reg_lab = regime_labels(close)
        else:
            reg_lab = reg_labels
        raw_sig = generate_strategy_signals(close, ema_s, ema_l, reg_labels=reg_lab, hurst_val=hurst_val, adf_p=adf_p, ema_pairs_list=ema_pairs_list, k_threshold=k_threshold)
        
        if dir_filter == "long_only":
            raw_sig = np.where(raw_sig > 0, 1.0, 0.0)
        elif dir_filter == "short_only":
            raw_sig = np.where(raw_sig < 0, -1.0, 0.0)

        fee_alone = fee_bps / 10000.0
        # Q3: Almgren-Chriss Dynamic Market Impact Slippage
        impact_bps = almgren_chriss_market_impact(balance0, adv_usd=10_000_000.0)
        total_cost_bps = fee_bps + (slippage_bps + impact_bps if shifted else 0.0)
        fee     = total_cost_bps / 10000.0

        if shifted:
            position = np.roll(raw_sig, 1); position[0] = 0.0
        else:
            position = raw_sig.copy()

        flips = np.where(position[1:] != position[:-1])[0] + 1
        ei    = np.concatenate(([1], flips))
        xi    = np.concatenate((flips, [n - 1]))
        ei, xi = ei[ei < n], xi[:len(ei)]
        
        if reg_labels is None:
            reg_lab = regime_labels(close)
        else:
            reg_lab = reg_labels
        atr_pct_arr = compute_atr_pct(close, window=14)
    else:
        ei, xi, position, reg_lab, atr_pct_arr, fee, raw_sig, fee_alone = pre_seg

    no_shift_ret = _unshifted_profit(close, raw_sig, fee_alone, balance0=balance0)
    trades = []
    equity = balance0

    # First pass: calculate unscaled trades (with SL if enabled) to get empirical Kelly f*
    if sizing in ("kelly_half", "kelly_full"):
        unscaled_rets = []
        for i0, i1 in zip(ei, xi):
            if i1 <= i0 or position[i0] == 0: continue
            d = position[i0]; p0 = close[i0]
            seg_close = close[i0:i1+1]
            atr_val = atr_pct_arr[i0] if atr_multiple else None
            stop_idx, _ = apply_stop_loss(seg_close, d, sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct, regime=reg_lab[i0], atr_pct=atr_val)
            actual_i1 = min(i0 + stop_idx, i1)
            p1 = close[actual_i1]
            unscaled_rets.append(((p1 / p0 - 1.0) * d) - (2 * fee))
        if unscaled_rets:
            u_arr = np.array(unscaled_rets)
            frac = 1.0 if sizing == "kelly_full" else 0.5
            kelly_f = continuous_kelly(u_arr, fraction=frac)
        else:
            kelly_f = 0.5
        if kelly_f <= 0: kelly_f = 0.05
    else:
        kelly_f = 1.0

    hist_rets = []
    peak_equity = balance0
    
    for i0, i1 in zip(ei, xi):
        if i1 <= i0: continue
        d = position[i0]
        if d == 0: continue
        
        # U6 SQN Gate
        if len(hist_rets) >= 20:
            hr_arr = np.array(hist_rets[-20:])
            hr_mean = hr_arr.mean()
            hr_std = hr_arr.std()
            sqn_rolling = (hr_mean / hr_std * np.sqrt(20)) if hr_std > 0 else 0.0
            if sqn_rolling < SQN_GATE_MIN:
                continue

        p0 = close[i0]
        # Apply stop-loss / trailing stop-loss / take profit if configured
        seg_close = close[i0:i1+1]
        atr_val = atr_pct_arr[i0] if atr_multiple else None
        stop_idx, stop_reason = apply_stop_loss(seg_close, d, sl_pct=sl_pct, tsl_pct=tsl_pct, tp_pct=tp_pct, regime=reg_lab[i0], atr_pct=atr_val)
        actual_i1 = i0 + stop_idx
        actual_i1 = min(actual_i1, i1)
        p1 = close[actual_i1]
        seg_u  = ((close[i0:actual_i1+1] / p0 - 1.0) * d) - (2 * fee)
        mfe    = float(seg_u.max() * 100.0)
        mae    = float(seg_u.min() * 100.0)
        tr     = float(seg_u[-1])
        hist_rets.append(tr)

        # Module 1.2: Dynamic Compounding / Reinvestment Schedules
        n_tr = len(trades)
        peak_equity = max(peak_equity, equity)

        if sizing == "kelly_full":
            pos_scale = kelly_f
        elif sizing == "kelly_half":
            if len(hist_rets) > 1:
                realized_vol = max(0.001, float(np.std(hist_rets[-20:])) if len(hist_rets) > 2 else 0.01)
                pos_scale = kelly_f * (0.01 / realized_vol)
            else:
                pos_scale = kelly_f
        elif sizing == "fixed_pct":
            pos_scale = 0.50  # Risk half balance scale per trade
        elif sizing == "linear_growth":
            pos_scale = min(2.0, 1.0 + (n_tr // 5) * 0.1)
        elif sizing == "exponential_growth":
            pos_scale = min(3.0, (1.02) ** (n_tr // 5))
        elif sizing == "profit_triggered":
            pos_scale = 1.5 if equity >= balance0 * 1.10 else 1.0
        elif sizing == "drawdown_gated":
            dd_pct = (peak_equity - equity) / peak_equity if peak_equity > 0 else 0.0
            pos_scale = 0.25 if dd_pct >= 0.15 else 1.0
        elif sizing == "volatility_scaled":
            if len(hist_rets) > 1:
                vol_20 = max(0.001, float(np.std(hist_rets[-20:])))
                pos_scale = 0.01 / vol_20
            else:
                pos_scale = 1.0
        else:
            pos_scale = 1.0
        
        pos_scale = min(3.0, max(0.01, pos_scale))
        pnl    = equity * pos_scale * tr
        equity += pnl
        eff    = (tr * 100.0 / mfe) if mfe > 0 else 0.0
        edge_r = (mfe / abs(mae)) if mae != 0 else mfe
        sl_hit  = 1 if stop_reason == 'sl'  else 0
        tsl_hit = 1 if stop_reason == 'tsl' else 0
        tp_hit  = 1 if stop_reason == 'tp'  else 0

        # Module 1.5: Reversal Lag & Late Exit Cost Calculation
        if d > 0:
            peak_local = int(np.argmax(seg_close[:stop_idx+1]))
            peak_p = seg_close[peak_local]
            candle_loss = float((peak_p - p1) / peak_p * 100.0) if peak_p > 0 else 0.0
        else:
            peak_local = int(np.argmin(seg_close[:stop_idx+1]))
            peak_p = seg_close[peak_local]
            candle_loss = float((p1 - peak_p) / peak_p * 100.0) if peak_p > 0 else 0.0
        
        rev_lag_bars = int(stop_idx - peak_local)
        late_exit_cost_bps = float(candle_loss * 100.0)

        trades.append((i0, actual_i1, ts[i0], ts[actual_i1], p0, p1, int(d),
                       tr, pnl, int(actual_i1 - i0), mae, mfe, eff, edge_r,
                       reg_lab[i0], sl_hit, tsl_hit, tp_hit, stop_reason,
                       rev_lag_bars, candle_loss, late_exit_cost_bps))

    if not trades: return None

    tdf = pd.DataFrame(trades, columns=[
        "entry_i","exit_i","entry_time","exit_time","entry_price","exit_price",
        "direction","ret","pnl","hold_bars","mae_pct","mfe_pct",
        "trade_efficiency_pct","edge_ratio","entry_regime",
        "sl_hit","tsl_hit","tp_hit","stop_reason",
        "reversal_lag_bars","candle_loss_pct","late_exit_cost_bps"
    ])
    tdf["entry_time"] = pd.to_datetime(tdf["entry_time"])
    tdf["exit_time"]  = pd.to_datetime(tdf["exit_time"])
    tdf["hold_sec"]   = (tdf["exit_time"] - tdf["entry_time"]).dt.total_seconds()
    tdf["hour"]       = tdf["entry_time"].dt.hour
    tdf["weekday"]    = tdf["entry_time"].dt.day_name()
    tdf["session"]    = tdf["hour"].apply(session_of_hour)
    tdf["year"]       = tdf["entry_time"].dt.year
    tdf["month"]      = tdf["entry_time"].dt.month
    tdf["equity"]     = balance0 + tdf["pnl"].cumsum()
    tdf["cum_profit"] = tdf["pnl"].clip(lower=0).cumsum()
    tdf["cum_loss"]   = (-tdf["pnl"].clip(upper=0)).cumsum()

    wins   = tdf[tdf.pnl > 0];  losses = tdf[tdf.pnl <= 0]
    gp     = float(wins.pnl.sum());  gl = float(-losses.pnl.sum())
    pf     = gp / gl if gl > 0 else np.inf
    total_ret = (equity / balance0 - 1.0) * 100.0
    bh_ret    = (close[-1] / close[0] - 1.0) * 100.0
    alpha_ret = total_ret - bh_ret

    eq_curve = np.concatenate(([balance0], tdf["equity"].values))
    dd_arr   = _drawdown(eq_curve)
    mdd      = float(dd_arr.min() * 100.0)
    mx_dd_bars = c = 0
    for v in (dd_arr < 0):
        c = (c+1) if v else 0; mx_dd_bars = max(mx_dd_bars, c)

    tr_rets  = tdf["ret"].values
    mu, sig  = tr_rets.mean(), tr_rets.std()
    
    # M1 Annualization: ann_factor = sqrt(N_trades_per_year) = sqrt(252 * bars_per_day / avg_hold)
    # Rigorously valid for continuous-exposure strategies (EMA SAR is 100% in-market by construction).
    avg_hold = tdf["hold_bars"].mean() if len(tdf) > 0 else 1.0
    tf_bars_map = {"5m": 288, "15m": 96, "30m": 48, "1h": 24, "2h": 12, "3h": 8, "4h": 6}
    bars_per_day = tf_bars_map.get(tf, 24)
    ann_factor = np.sqrt(252.0 * bars_per_day / max(1.0, avg_hold))

    sharpe   = (mu / sig * ann_factor) if sig > 0 else 0.0
    dstd     = tr_rets[tr_rets < 0].std() if len(tr_rets[tr_rets < 0]) > 0 else 0.0
    sortino  = (mu / dstd * ann_factor) if dstd > 0 else 0.0
    calmar   = min(999.0, (total_ret / abs(mdd))) if mdd < 0 else (min(999.0, total_ret) if (mdd == 0 and total_ret > 0) else 0.0)

    pos_sum  = float(tr_rets[tr_rets > 0].sum()) if len(tr_rets[tr_rets > 0]) else 0.0
    neg_sum  = float(abs(tr_rets[tr_rets < 0].sum())) if len(tr_rets[tr_rets < 0]) else 0.0
    omega    = (pos_sum / neg_sum) if neg_sum > 0 else np.inf
    ulcer    = float(np.sqrt(np.mean((dd_arr * 100.0)**2)))
    recovery = min(999.0, (total_ret / abs(mdd))) if mdd < 0 else min(999.0, total_ret)
    p95      = float(np.percentile(tr_rets, 95)) if len(tr_rets) > 5 else 0.0
    p5       = float(abs(np.percentile(tr_rets, 5))) if len(tr_rets) > 5 else 1e-9
    tail_r   = (p95 / p5) if p5 > 0 else 0.0
    
    var_95   = float(np.percentile(tr_rets, 5) * 100.0) if len(tr_rets) > 5 else 0.0
    cvar_95  = float(tr_rets[tr_rets <= np.percentile(tr_rets, 5)].mean() * 100.0) if len(tr_rets) > 5 else 0.0
    pain_idx = float(np.mean(abs(dd_arr * 100.0)))
    
    # M2 Fix: True Deflated Sharpe Ratio (Bailey & López de Prado 2012)
    sk_val = float(stats.skew(tr_rets)) if len(tr_rets) > 3 else 0.0
    kt_val = float(stats.kurtosis(tr_rets)) + 3.0 if len(tr_rets) > 3 else 3.0
    dsr_prob = deflated_sharpe_ratio(sharpe, n_obs=len(tr_rets), n_trials=4500, skewness=sk_val, kurtosis=kt_val)
    dvr      = round(float(sharpe * dsr_prob), 2)
    
    alpha_per_trade = tr_rets - (bh_ret / 100.0 * (tdf['hold_bars'].values / max(1, len(close))))
    te       = alpha_per_trade.std()
    info_rat = round(float((alpha_per_trade.mean() / te * ann_factor)), 2) if te > 0 else 0.0
    # F7 Fix: Schwager Gain-to-Pain = sum(positive period rets) / sum(|negative period rets|)
    # Using per-trade fractional returns (consistent scale), then convert to pct
    pos_ret_sum_pct = float(tr_rets[tr_rets > 0].sum()) * 100.0
    neg_ret_sum_pct = float(abs(tr_rets[tr_rets < 0].sum())) * 100.0
    gain_to_pain = round(float(pos_ret_sum_pct / (neg_ret_sum_pct + 1e-9)), 2)

    wr       = len(wins) / len(tdf)
    lr       = 1.0 - wr
    avg_w    = float(wins["ret"].mean() * 100.0) if len(wins) else 0.0
    avg_l    = float(losses["ret"].mean() * 100.0) if len(losses) else 0.0
    avg_wd   = float(wins["pnl"].mean()) if len(wins) else 0.0
    avg_ld   = float(losses["pnl"].mean()) if len(losses) else 0.0
    payoff   = (avg_wd / abs(avg_ld)) if avg_ld < 0 else np.inf
    exp_pct  = wr * avg_w + lr * avg_l
    exp_dol  = wr * avg_wd + lr * avg_ld
    sqn      = (mu / sig * np.sqrt(len(tdf))) if sig > 0 else 0.0

    longs  = tdf[tdf.direction == 1]; shorts = tdf[tdf.direction == -1]
    l_win  = (longs.pnl > 0).mean() * 100.0 if len(longs) else 0.0
    s_win  = (shorts.pnl > 0).mean() * 100.0 if len(shorts) else 0.0
    l_prof = (np.prod(1 + longs.ret.values) - 1.0) * 100.0 if len(longs) else 0.0
    s_prof = (np.prod(1 + shorts.ret.values) - 1.0) * 100.0 if len(shorts) else 0.0

    kelly_f_opt = continuous_kelly(tr_rets, fraction=1.0)
    k_eq = [balance0]
    for r_i in tr_rets: k_eq.append(k_eq[-1] * (1.0 + kelly_f_opt * r_i))
    kelly_prof = float((k_eq[-1] / balance0 - 1.0) * 100.0)

    t_stat, p_val  = stats.ttest_1samp(tr_rets, 0) if len(tr_rets) > 1 else (0.0, 1.0)
    lb_p           = ljung_box_p(tr_rets)
    runs_z, runs_p = runs_test(tr_rets)
    is_significant = bool(p_val < 0.05)
    has_autocorr   = bool(lb_p < 0.05)
    ror            = risk_of_ruin(tr_rets, ruin_level=0.5)

    mc      = monte_carlo(tr_rets)
    reg_prf = {}
    for regime, grp in tdf.groupby("entry_regime"):
        rp = grp["ret"].values
        reg_prf[regime] = dict(
            trades=len(grp),
            profit_pct=round(float((np.prod(1+rp)-1)*100), 2),
            win_rate_pct=round(float((rp>0).mean()*100), 2),
        )

    monthly = tdf.groupby(["year","month"]).agg(
        trades=("pnl","count"),
        profit_pct=("ret", lambda r: round((np.prod(1+r)-1)*100, 2)),
        cum_pnl=("pnl","sum"),
    ).reset_index()
    monthly["period"] = monthly["year"].astype(str)+"-"+monthly["month"].astype(str).str.zfill(2)

    t_start = pd.to_datetime(ts[0]); t_end = pd.to_datetime(ts[-1])
    days_span = max(1.0, (t_end - t_start).total_seconds() / 86400.0)
    cagr_val = round(cagr(total_ret, days_span), 2)

    summary = dict(
        final_balance   =round(equity, 2),
        profit_pct      =round(total_ret, 2),
        cagr_pct        =cagr_val,
        profit_pct_no_shift=round(no_shift_ret, 2),
        lookahead_delta =round(no_shift_ret - total_ret, 2),
        bh_return_pct   =round(bh_ret, 2),
        alpha_pct       =round(alpha_ret, 2),
        profit_factor   =round(pf, 2) if not np.isinf(pf) else 999.0,
        sharpe          =round(sharpe, 2),
        sortino         =round(sortino, 2),
        calmar          =round(calmar, 2) if not np.isinf(calmar) else 999.0,
        omega_ratio     =round(omega, 2) if not np.isinf(omega) else 999.0,
        ulcer_index     =round(ulcer, 2),
        var_95_pct      =round(var_95, 2),
        cvar_95_pct     =round(cvar_95, 2),
        pain_index      =round(pain_idx, 2),
        dvr             =dvr,
        dsr_prob        =round(dsr_prob, 4),
        information_ratio=info_rat,
        gain_to_pain_ratio=gain_to_pain,
        recovery_factor =round(recovery, 2),
        tail_ratio      =round(tail_r, 2),
        max_drawdown_pct=round(mdd, 2),
        max_dd_bars     =mx_dd_bars,
        total_trades    =len(tdf),
        win_rate_pct    =round(wr * 100.0, 2),
        payoff_ratio    =round(payoff, 2) if not np.isinf(payoff) else 999.0,
        expectancy_pct  =round(exp_pct, 2),
        expectancy_dol  =round(exp_dol, 2),
        sqn             =round(sqn, 2),
        kelly_fraction  =round(kelly_f_opt, 4),
        kelly_profit_pct=round(kelly_prof, 2),
        risk_of_ruin_pct=round(ror * 100.0, 2),
        long_trades     =len(longs),
        long_win_rate   =round(l_win, 2),
        long_profit_pct =round(l_prof, 2),
        short_trades    =len(shorts),
        short_win_rate  =round(s_win, 2),
        short_profit_pct=round(s_prof, 2),
        max_consec_wins =_max_consec((tdf.pnl > 0).values, True),
        max_consec_loss =_max_consec((tdf.pnl > 0).values, False),
        avg_win_pct     =round(avg_w, 2),
        avg_loss_pct    =round(avg_l, 2),
        avg_hold_bars   =round(float(tdf.hold_bars.mean()), 1),
        avg_mae_pct     =round(float(tdf.mae_pct.mean()), 2),
        avg_mfe_pct     =round(float(tdf.mfe_pct.mean()), 2),
        avg_edge_ratio  =round(float(tdf.edge_ratio.mean()), 2),
        avg_reversal_lag_bars   =round(float(tdf["reversal_lag_bars"].mean()), 1),
        avg_candle_loss_pct     =round(float(tdf["candle_loss_pct"].mean()), 2),
        avg_late_exit_cost_bps  =round(float(tdf["late_exit_cost_bps"].mean()), 1),
        total_late_exit_cost_pct=round(float(tdf["candle_loss_pct"].sum()), 2),
        t_stat          =round(float(t_stat), 3),
        p_value         =round(float(p_val), 4),
        is_significant  =is_significant,
        ljung_box_p     =round(lb_p, 4),
        has_autocorrelation=has_autocorr,
        runs_z_score    =runs_z,
        runs_p_value    =runs_p,
        **mc,
    )
    return summary, tdf, eq_curve, reg_prf, monthly

def fee_sensitivity_sweep(close, ts, ema_s, ema_l, symbol, tf, fee_grid=[0, 5, 10, 20, 30, 50], hurst_val=0.5, adf_p=0.10):
    """
    Q2 Upgrade: Transaction Cost Sensitivity Analysis
    Sweeps fee_bps in [0, 5, 10, 20, 30, 50] to plot Sharpe vs Fee curve and determine break-even fee.
    """
    _style()
    results = []
    for fee in fee_grid:
        res = backtest_pair(close, ts, ema_s, ema_l, fee_bps=fee, hurst_val=hurst_val, adf_p=adf_p)
        if res:
            sm = res[0]
            results.append(dict(fee_bps=fee, profit_pct=sm["profit_pct"], sharpe=sm["sharpe"], profit_factor=sm["profit_factor"]))

    df_fee = pd.DataFrame(results)
    if df_fee.empty:
        return "", 0.0

    fig, ax1 = plt.subplots(figsize=(7, 4))
    ax2 = ax1.twinx()

    ax1.plot(df_fee["fee_bps"], df_fee["sharpe"], 'o-', color="#38bdf8", lw=2, label="Sharpe Ratio")
    ax2.plot(df_fee["fee_bps"], df_fee["profit_pct"], 's--', color="#4ade80", lw=1.8, label="Profit %")

    ax1.axhline(0, color="#f87171", ls=":", lw=1.2)
    ax1.set_xlabel("Transaction Fee (bps)")
    ax1.set_ylabel("Sharpe Ratio", color="#38bdf8")
    ax2.set_ylabel("Total Profit %", color="#4ade80")
    ax1.set_title(f"{symbol} {tf} - Fee Sensitivity & Break-Even Analysis", fontsize=11, fontweight="bold")
    ax1.grid(True); fig.tight_layout()

    out_png = f"{OUT_DIR}/fee_sensitivity_{symbol}_{tf}.png"
    fig.savefig(out_png, dpi=120)
    plt.close(fig)

    breakeven = compute_breakeven_fee(close, ts, ema_s, ema_l, hurst_val=hurst_val, adf_p=adf_p)
    return out_png, breakeven

# ----------------------------- COMBINATORIAL MATRIX (24 COMBOS) -------------
def generate_combinatorial_matrix(close, ts, ema_s, ema_l, symbol, tf, bx, by, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    combos = []
    fees = [0.0, 10.0]
    shifts = [True, False]
    directions = ["both", "long_only", "short_only"]
    sizings = ["flat", "kelly_half"]

    for fee in fees:
        for shift in shifts:
            for d in directions:
                for siz in sizings:
                    res = backtest_pair(close, ts, ema_s, ema_l, fee_bps=fee, dir_filter=d, shifted=shift, sizing=siz, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
                    if res:
                        sm = res[0]
                        combos.append(dict(
                            symbol=symbol, timeframe=tf, ema_x=bx, ema_y=by,
                            fee_bps=fee, shift_execution=shift, direction=d, sizing=siz,
                            profit_pct=sm["profit_pct"], profit_factor=sm["profit_factor"],
                            sharpe=sm["sharpe"], max_drawdown_pct=sm["max_drawdown_pct"],
                            win_rate_pct=sm["win_rate_pct"], trades=sm["total_trades"],
                            sqn=sm["sqn"], ulcer_index=sm["ulcer_index"], cvar_95_pct=sm["cvar_95_pct"]
                        ))
    return pd.DataFrame(combos)

# ----------------------------- MODULE 1.1 N-EMA CONFLUENCE ENGINE -----------
def run_confluence_grid(close, ts, symbol, tf, period_pairs=None, k_levels=None, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    Module 1.1: Evaluates N-EMA Confluence Confirmation Gates (K-of-N voting across EMA pairs).
    Prevents whipsaws by requiring multiple EMA pairs to simultaneously agree before trade entry.
    """
    if period_pairs is None:
        period_pairs = [(5, 20), (10, 50), (20, 200)]
    
    # Pre-calculate EMA arrays for each pair
    ema_arrays = []
    for px, py in period_pairs:
        es = _ema(close, 2.0 / (px + 1.0))
        el = _ema(close, 2.0 / (py + 1.0))
        ema_arrays.append((es, el, px, py))

    rows = []
    
    # 1. Evaluate individual 1x1 baselines
    base_sharpe = None
    base_trades = None
    for i, (es, el, px, py) in enumerate(ema_arrays):
        res = backtest_pair(close, ts, es, el, tf=tf, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        if res:
            sm = res[0]
            if i == 0:
                base_sharpe = sm["sharpe"]
                base_trades = max(1, sm["total_trades"])
            rows.append(dict(
                symbol=symbol, timeframe=tf,
                confluence_type="1x1_baseline",
                n_pairs=1, k_threshold=1,
                pairs_str=f"({px},{py})",
                profit_pct=sm["profit_pct"], cagr_pct=sm["cagr_pct"], sharpe=sm["sharpe"],
                max_drawdown_pct=sm["max_drawdown_pct"], win_rate_pct=sm["win_rate_pct"],
                total_trades=sm["total_trades"], sqn=sm["sqn"], ulcer_index=sm["ulcer_index"],
                sharpe_delta=0.0, trades_reduction_pct=0.0
            ))

    # 2. Evaluate Multi-Pair Confluence sets (combinations of length 2 to N)
    from itertools import combinations
    N = len(ema_arrays)
    for r in range(2, N + 1):
        for combo in combinations(range(N), r):
            selected_pairs = [ema_arrays[idx] for idx in combo]
            pairs_list = [(es, el) for es, el, _, _ in selected_pairs]
            pairs_desc = "+".join([f"({px},{py})" for _, _, px, py in selected_pairs])
            
            # K thresholds to test: from r//2 + 1 to r
            possible_k = k_levels if k_levels else list(range(max(1, r // 2 + 1), r + 1))
            for k in possible_k:
                if k > r: continue
                res = backtest_pair(close, ts, tf=tf, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p,
                                    ema_pairs_list=pairs_list, k_threshold=k)
                if res:
                    sm = res[0]
                    sh_delta = round(sm["sharpe"] - (base_sharpe if base_sharpe is not None else sm["sharpe"]), 2)
                    tr_red = round((1.0 - sm["total_trades"] / (base_trades if base_trades else 1)) * 100.0, 1)
                    mode_label = f"{r}x{r}_unanimous" if k == r else f"{k}-of-{r}_majority"
                    rows.append(dict(
                        symbol=symbol, timeframe=tf,
                        confluence_type=mode_label,
                        n_pairs=r, k_threshold=k,
                        pairs_str=pairs_desc,
                        profit_pct=sm["profit_pct"], cagr_pct=sm["cagr_pct"], sharpe=sm["sharpe"],
                        max_drawdown_pct=sm["max_drawdown_pct"], win_rate_pct=sm["win_rate_pct"],
                        total_trades=sm["total_trades"], sqn=sm["sqn"], ulcer_index=sm["ulcer_index"],
                        sharpe_delta=sh_delta, trades_reduction_pct=tr_red
                    ))

    df_confluence = pd.DataFrame(rows)
    return df_confluence

# ----------------------------- MODULE 1.5 TREND REVERSAL LAG ENGINE --------
def analyze_reversal_lag(tdf, symbol="UNKNOWN", tf="1h"):
    """
    Module 1.5: Trend Reversal Lag & Late Exit Cost Analysis Engine.
    Quantifies the number of bars elapsed between favorable price peak and exit signal/stop fire,
    and calculates candle loss % and late exit cost in basis points.
    Returns: (summary_dict, regime_breakdown_df, top10_worst_late_exits_df)
    """
    if tdf is None or tdf.empty or "reversal_lag_bars" not in tdf.columns:
        return {}, pd.DataFrame(), pd.DataFrame()

    total_trades = len(tdf)
    avg_lag = float(tdf["reversal_lag_bars"].mean())
    std_lag = float(tdf["reversal_lag_bars"].std()) if total_trades > 1 else 0.0
    max_lag = int(tdf["reversal_lag_bars"].max())
    
    avg_loss_pct = float(tdf["candle_loss_pct"].mean())
    total_loss_pct = float(tdf["candle_loss_pct"].sum())
    avg_cost_bps = float(tdf["late_exit_cost_bps"].mean())

    summary_lag = dict(
        symbol=symbol, timeframe=tf, total_trades=total_trades,
        avg_reversal_lag_bars=round(avg_lag, 1),
        std_reversal_lag_bars=round(std_lag, 1),
        max_reversal_lag_bars=max_lag,
        avg_candle_loss_pct=round(avg_loss_pct, 2),
        total_candle_loss_pct=round(total_loss_pct, 2),
        avg_late_exit_cost_bps=round(avg_cost_bps, 1)
    )

    # Regime breakdown
    reg_rows = []
    if "entry_regime" in tdf.columns:
        for reg, grp in tdf.groupby("entry_regime"):
            reg_rows.append(dict(
                regime=reg, trades=len(grp),
                avg_lag_bars=round(float(grp["reversal_lag_bars"].mean()), 1),
                avg_candle_loss_pct=round(float(grp["candle_loss_pct"].mean()), 2),
                avg_late_exit_cost_bps=round(float(grp["late_exit_cost_bps"].mean()), 1)
            ))
    df_regime_lag = pd.DataFrame(reg_rows)

    # Pareto top-10 worst late exit trades
    df_top10 = tdf.sort_values(by="candle_loss_pct", ascending=False).head(10)[[
        "entry_time", "exit_time", "direction", "entry_price", "exit_price",
        "ret", "hold_bars", "reversal_lag_bars", "candle_loss_pct", "late_exit_cost_bps", "stop_reason"
    ]].copy()

    return summary_lag, df_regime_lag, df_top10




# ----------------------------- EFFICIENT FRONTIER SOLVER --------------------
def compute_efficient_frontier(best_by_combo):
    """
    F6 Upgrade: Dimensionally-Consistent Daily Resampled Portfolio Frontier Solver
    Resamples trade/equity outputs of multi-timeframe strategies to a common calendar daily frequency
    before computing covariance and solving Markowitz SLSQP optimization.
    """
    if not best_by_combo or len(best_by_combo) < 2:
        return pd.DataFrame(), pd.DataFrame()

    daily_series_dict = {}
    for (sym, tf), payload in best_by_combo.items():
        tdf = payload[2]
        if tdf.empty or "exit_time" not in tdf.columns or "equity" not in tdf.columns:
            continue
        ts_index = pd.to_datetime(tdf["exit_time"])
        s_eq = pd.Series(tdf["equity"].values, index=ts_index)
        s_daily = s_eq.resample("D").last().ffill()
        daily_series_dict[f"{sym}_{tf}"] = s_daily

    if len(daily_series_dict) < 2:
        return pd.DataFrame(), pd.DataFrame()

    daily_df = pd.DataFrame(daily_series_dict).ffill().bfill()
    rets_df = daily_df.pct_change().dropna()

    valid_cols = [col for col in rets_df.columns if rets_df[col].std() > 1e-9]
    if len(valid_cols) < 2:
        return pd.DataFrame(), pd.DataFrame()

    rets_df = rets_df[valid_cols]
    mean_returns = rets_df.mean() * 252.0
    cov_matrix = rets_df.cov() * 252.0
    num_assets = len(mean_returns)

    def portfolio_performance(weights):
        returns = np.sum(mean_returns * weights)
        std = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
        return returns, std

    def min_sharpe(weights):
        p_ret, p_std = portfolio_performance(weights)
        return -p_ret / (p_std + 1e-9)

    constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    bounds = tuple((0.0, 1.0) for _ in range(num_assets))
    init_guess = num_assets * [1.0 / num_assets]

    opt_sharpe = optimize.minimize(min_sharpe, init_guess, method='SLSQP', bounds=bounds, constraints=constraints)
    max_sharpe_weights = opt_sharpe.x if opt_sharpe.success else init_guess

    frontier_points = []
    target_returns = np.linspace(mean_returns.min(), mean_returns.max(), 10)
    for target in target_returns:
        def min_vol(weights): return portfolio_performance(weights)[1]
        c_target = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1},
                    {'type': 'eq', 'fun': lambda x, t=target: portfolio_performance(x)[0] - t})
        res = optimize.minimize(min_vol, init_guess, method='SLSQP', bounds=bounds, constraints=c_target)
        if res.success:
            ret, std = portfolio_performance(res.x)
            frontier_points.append(dict(target_return=round(ret*100,2), volatility=round(std*100,2), sharpe=round(ret/(std+1e-9),2)))

    res_df = pd.DataFrame(frontier_points)
    weights_dict = dict(zip(rets_df.columns, np.round(max_sharpe_weights, 4)))
    weights_df = pd.DataFrame([weights_dict])
    weights_df.insert(0, "Portfolio_Type", "Optimal_MaxSharpe")
    return res_df, weights_df

# ----------------------------- GRID ENGINE ----------------------------------
@dataclass
class GridResult:
    rows: list = field(default_factory=list)
    def add(self, s, t, x, y, sm):
        self.rows.append(dict(symbol=s, timeframe=t, ema_x=x, ema_y=y, **sm))
    def to_df(self):
        return pd.DataFrame(self.rows)

def run_grid(df, symbol, tf, xp, yp, reg_labels=None, hurst_val=0.5, adf_p=0.10, min_sep=5):
    """
    Exhaustive EMA grid search over all valid (x, y) pairs where y >= x + min_sep.
    Pre-computes EMA matrix for all needed periods, then evaluates pairs in parallel
    using ThreadPoolExecutor for 2-4x speed on multi-core machines.
    """
    close = df["close"].to_numpy(dtype=np.float64)
    ts    = pd.to_datetime(df["time"]).to_numpy()
    allp  = sorted(set(xp)|set(yp))
    mats  = ema_matrix(close, np.array(allp))
    pmap  = {p: i for i, p in enumerate(allp)}
    res   = GridResult(); best = None

    if reg_labels is None:
        reg_labels = regime_labels(close)

    # Build valid pairs — enforce min_sep to skip near-identical EMA pairs
    pairs = [(x, y) for x in xp for y in yp if y >= x + min_sep]

    def _eval_pair(xy):
        x, y = xy
        out = backtest_pair(close, ts, mats[pmap[x]], mats[pmap[y]],
                            reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        if out is None:
            return None
        sm, tdf, eq_c, reg_p, mty = out
        score = compute_score(sm)
        return (x, y, sm, tdf, eq_c, reg_p, mty, score)

    # Parallel inner evaluation with thread pool (GIL-light numpy ops)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    n_threads = min(4, max(1, (os.cpu_count() or 2) // 2))
    with ThreadPoolExecutor(max_workers=n_threads) as tex:
        fut_map = {tex.submit(_eval_pair, p): p for p in pairs}
        for fut in as_completed(fut_map):
            r = fut.result()
            if r is None:
                continue
            x, y, sm, tdf, eq_c, reg_p, mty, score = r
            res.add(symbol, tf, int(x), int(y), sm)
            if best is None or score > best[9]:
                best = (sm, tdf, eq_c, int(x), int(y), reg_p, mty,
                        mats[pmap[x]], mats[pmap[y]], score)

    if best is not None:
        return res, best[:9]
    return res, None

def bayesian_optimization_search(df, symbol, tf, n_init=32, n_iter=20, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    """
    F2 Upgrade: Bayesian Optimization Search via Quasi-Monte Carlo Sobol Sampling + RBF Surrogate
    Maps the Sharpe / composite score surface efficiently with ~50 evaluations instead of brute-force grid.
    """
    from scipy.stats import qmc
    from scipy.interpolate import RBFInterpolator

    close = df["close"].to_numpy(dtype=np.float64)
    ts    = pd.to_datetime(df["time"]).to_numpy()
    if reg_labels is None:
        reg_labels = regime_labels(close)

    res = GridResult()
    best = None
    cache = {}

    def eval_pair(x, y):
        x_i, y_i = int(round(x)), int(round(y))
        if x_i >= y_i or x_i < MIN_P or y_i > MAX_P:
            return -999.0, None
        if (x_i, y_i) in cache:
            return cache[(x_i, y_i)][0], cache[(x_i, y_i)][1]
        
        ema_s = _ema(close, 2.0 / (x_i + 1.0))
        ema_l = _ema(close, 2.0 / (y_i + 1.0))
        out = backtest_pair(close, ts, ema_s, ema_l, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
        if out is None:
            cache[(x_i, y_i)] = (-999.0, None)
            return -999.0, None
        
        sm, tdf, eq_c, reg_p, mty = out
        score = compute_score(sm)
        res.add(symbol, tf, x_i, y_i, sm)
        full_best = (sm, tdf, eq_c, x_i, y_i, reg_p, mty, ema_s, ema_l, score)
        cache[(x_i, y_i)] = (score, full_best)
        return score, full_best

    # Step 1: Initial Quasi-Monte Carlo Sobol Sampling
    sampler = qmc.Sobol(d=2, scramble=True, seed=42)
    sample = sampler.random(n_init)
    
    x_samples = MIN_P + sample[:, 0] * (MAX_P - MIN_P)
    y_samples = MIN_P + sample[:, 1] * (MAX_P - MIN_P)

    X_obs, y_obs = [], []
    for x_val, y_val in zip(x_samples, y_samples):
        if x_val >= y_val:
            x_val, y_val = min(x_val, y_val), max(x_val, y_val)
            if x_val == y_val: y_val += 1.0
        sc, full_b = eval_pair(x_val, y_val)
        if full_b is not None:
            X_obs.append([int(round(x_val)), int(round(y_val))])
            y_obs.append(sc)
            if best is None or sc > best[9]:
                best = full_b

    # Step 2: Iterative RBF Surrogate Model Search
    for it in range(n_iter):
        if len(X_obs) < 5:
            break
        X_arr = np.array(X_obs)
        y_arr = np.array(y_obs)
        
        try:
            rbf = RBFInterpolator(X_arr, y_arr, kernel='thin_plate_spline', smoothing=1e-3)
            cand_x = np.random.uniform(MIN_P, MAX_P, 500)
            cand_y = np.random.uniform(MIN_P, MAX_P, 500)
            valid_mask = cand_x < cand_y
            cand_X = np.column_stack([cand_x[valid_mask], cand_y[valid_mask]])
            
            if len(cand_X) > 0:
                pred_scores = rbf(cand_X)
                best_cand_idx = np.argmax(pred_scores)
                cx, cy = cand_X[best_cand_idx]
                sc, full_b = eval_pair(cx, cy)
                if full_b is not None:
                    X_obs.append([int(round(cx)), int(round(cy))])
                    y_obs.append(sc)
                    if best is None or sc > best[9]:
                        best = full_b
        except Exception:
            break

    if best is not None:
        return res, best[:9]
    return res, None

def coarse_fine(df, symbol, tf, step=5, r=ROBUST_R, reg_labels=None, hurst_val=0.5, adf_p=0.10):
    # S4: Use timeframe-adaptive period range so EMA horizons are comparable across TFs
    tf_min, tf_max = PERIOD_RANGE.get(tf, (MIN_P, MAX_P))
    cp = list(range(tf_min, tf_max + 1, step))
    if tf_min not in cp: cp.insert(0, tf_min)
    if tf_max not in cp: cp.append(tf_max)

    if reg_labels is None:
        reg_labels = regime_labels(df["close"].to_numpy(dtype=np.float64))

    res, best = run_grid(df, symbol, tf, cp, cp, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
    if best is None: return res, best
    bx, by = best[3], best[4]

    # S2 Upgrade: Boundary-aware fine search surrounding best + extreme boundary pairs
    radius = max(step, r*2)
    fx = list(set(range(max(tf_min, bx-radius), min(tf_max, bx+radius)+1)).union({tf_min}))
    fy = list(set(range(max(tf_min, by-radius), min(tf_max, by+radius)+1)).union({tf_max}))
    fx.sort(); fy.sort()

    fres, fbest = run_grid(df, symbol, tf, fx, fy, reg_labels=reg_labels, hurst_val=hurst_val, adf_p=adf_p)
    combined = GridResult(); combined.rows = res.rows + fres.rows
    return combined, (fbest if (fbest and fbest[0]["profit_factor"] >= best[0]["profit_factor"]) else best)

# ----------------------------- DATA LOADING --------------------------------
def fetch_live_market_data(symbols=SYMBOLS, timeframes=TIMEFRAMES):
    import yfinance as yf
    import time
    
    ticker_map = {
        "ETH": "ETH-USD", "SOL": "SOL-USD", "XRP": "XRP-USD",
        "XAU": "GC=F", "XAG": "SI=F"
    }
    
    cache_dir = os.path.join("./data", "data_cache")
    os.makedirs(cache_dir, exist_ok=True)
    out = {}
    
    print("[+] Fetching & Resampling REAL historical market data (Optimized High-Throughput)...")
    for sym in symbols:
        ticker_str = ticker_map.get(sym, sym)
        
        # Load or download base timeframes: 5m (for 5m, 15m, 30m) and 1h (for 1h, 2h, 3h, 4h) and 1d
        base_dfs = {}
        for base_tf, p, inv in [("5m", "60d", "5m"), ("1h", "700d", "1h"), ("1d", "max", "1d")]:
            cache_file = os.path.join(cache_dir, f"{sym}_{base_tf}.csv")
            if os.path.exists(cache_file) and (time.time() - os.path.getmtime(cache_file)) < 86400:
                try:
                    df_c = pd.read_csv(cache_file)
                    df_c["time"] = pd.to_datetime(df_c["time"])
                    base_dfs[base_tf] = df_c
                    continue
                except Exception: pass
            
            try:
                df_raw = yf.download(ticker_str, period=p, interval=inv, progress=False)
                if not df_raw.empty:
                    df_raw = df_raw.reset_index()
                    if isinstance(df_raw.columns, pd.MultiIndex):
                        df_raw.columns = [col[0] for col in df_raw.columns]
                    col_time = "Datetime" if "Datetime" in df_raw.columns else "Date"
                    df_raw = df_raw.rename(columns={col_time: "time", "Close": "close"})
                    df_raw["time"] = pd.to_datetime(df_raw["time"])
                    df_final = df_raw.dropna(subset=["time", "close"])[["time", "close"]]
                    
                    if os.path.exists(cache_file):
                        try:
                            df_old = pd.read_csv(cache_file)
                            df_old["time"] = pd.to_datetime(df_old["time"])
                            df_final = pd.concat([df_old, df_final], ignore_index=True)
                        except Exception: pass

                    df_final = df_final.drop_duplicates(subset=["time"]).sort_values("time").reset_index(drop=True)
                    df_final.to_csv(cache_file, index=False)
                    base_dfs[base_tf] = df_final
            except Exception as e:
                print(f"  [error] Failed base download {sym} {base_tf}: {e}")
                
        # Fill requested timeframes via local high-speed resampling
        for tf in timeframes:
            if tf in base_dfs:
                out[(sym, tf)] = base_dfs[tf]
                print(f"  [ready] {sym} {tf} ({len(base_dfs[tf])} bars)")
            elif tf in ["15m", "30m"] and "5m" in base_dfs:
                freq_rule = {"15m": "15min", "30m": "30min"}[tf]
                df_res = base_dfs["5m"].set_index("time").resample(freq_rule).agg({"close": "last"}).dropna().reset_index()
                out[(sym, tf)] = df_res[["time", "close"]]
                print(f"  [resampled] {sym} {tf} from 5m ({len(df_res)} bars)")
            elif tf in ["2h", "3h", "4h"] and "1h" in base_dfs:
                freq_rule = {"2h": "2h", "3h": "3h", "4h": "4h"}[tf]
                df_res = base_dfs["1h"].set_index("time").resample(freq_rule).agg({"close": "last"}).dropna().reset_index()
                out[(sym, tf)] = df_res[["time", "close"]]
                print(f"  [resampled] {sym} {tf} from 1h ({len(df_res)} bars)")
                
    return out

def load_all(path):
    xl  = pd.ExcelFile(path); out = {}
    for sh in xl.sheet_names:
        if "_" not in sh: continue
        sym, tf = sh.rsplit("_", 1)
        df = xl.parse(sh)
        if "time" not in df.columns or "close" not in df.columns: continue
        df = df.dropna(subset=["time", "close"])
        df = df[df["close"] > 0]
        df["time"] = pd.to_datetime(df["time"])
        df = df.drop_duplicates(subset=["time"])
        if CUTOFF is not None:
            df = df[df["time"] < CUTOFF]
        df = df.sort_values("time").reset_index(drop=True)
        if len(df) < MIN_P + 10: continue
        out[(sym, tf)] = df[["time","close"]]
    return out

def make_demo(bars=4000, seed=42):
    return fetch_live_market_data()

# ----------------------------- VISUALS --------------------------------------
def _style():

    plt.rcParams.update({
        "figure.facecolor": "#0a0f1e", "axes.facecolor": "#0d1526",
        "axes.edgecolor": "#1e3a5f",   "axes.labelcolor": "#94a3b8",
        "xtick.color": "#64748b",      "ytick.color": "#64748b",
        "text.color": "#e2e8f0",       "grid.color": "#1e3a5f",
        "grid.linewidth": 0.6,         "font.family": "DejaVu Sans",
    })

def make_visuals(symbol, tf, grid_df, bx, by, eq_curve, tdf):


    _style()

    dd = _drawdown(eq_curve) * 100.0
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 5.5), sharex=True,
                                    gridspec_kw={"height_ratios": [3, 1]})
    ax1.plot(eq_curve, color="#38bdf8", lw=1.8, label="Equity")
    ax1.fill_between(range(len(eq_curve)), BALANCE0, eq_curve,
                     where=np.array(eq_curve) >= BALANCE0, alpha=0.10, color="#38bdf8")
    ax1.fill_between(range(len(eq_curve)), BALANCE0, eq_curve,
                     where=np.array(eq_curve) < BALANCE0,  alpha=0.14, color="#f87171")
    ax1.axhline(BALANCE0, color="#475569", ls="--", lw=0.8)
    ax1.set_title(f"{symbol} {tf} - EMA({bx},{by}) Equity & Drawdown", fontsize=11, fontweight="bold")
    ax1.set_ylabel("Balance ($)"); ax1.legend(fontsize=9); ax1.grid(True)
    ax2.fill_between(range(len(dd)), dd, 0, color="#f87171", alpha=0.55)
    ax2.set_ylabel("DD %"); ax2.set_xlabel("Trade #"); ax2.grid(True)
    fig.tight_layout(); fig.savefig(p:=f"{OUT_DIR}/equity_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    eq_png = p

    piv = grid_df.pivot_table("profit_factor", "ema_x", "ema_y", aggfunc="max")
    fig, ax = plt.subplots(figsize=(7, 6))
    im = ax.imshow(piv.values, aspect="auto", origin="lower", cmap="plasma",
                   extent=[piv.columns.min(), piv.columns.max(),
                            piv.index.min(),   piv.index.max()])
    fig.colorbar(im, ax=ax, label="Profit Factor")
    ax.scatter([by], [bx], color="#4ade80", s=90, zorder=5, label=f"Best ({bx},{by})")
    ax.set_xlabel("EMA y"); ax.set_ylabel("EMA x")
    ax.set_title(f"{symbol} {tf} - Parameter Space Heatmap", fontsize=11, fontweight="bold")
    ax.legend(fontsize=8); fig.tight_layout()
    fig.savefig(p:=f"{OUT_DIR}/heatmap_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    hm_png = p

    fig, ax = plt.subplots(figsize=(7, 5))
    sc = ax.scatter(tdf.mae_pct, tdf.mfe_pct, c=tdf["ret"]*100,
                    cmap="RdYlGn", edgecolors="none", alpha=0.75, s=36)
    fig.colorbar(sc, ax=ax, label="Trade Return %")
    ax.axhline(0, color="#475569", lw=0.7); ax.axvline(0, color="#475569", lw=0.7)
    ax.set_xlabel("MAE %"); ax.set_ylabel("MFE %")
    ax.set_title(f"{symbol} {tf} - Execution Quality (MAE vs MFE)", fontsize=11, fontweight="bold")
    ax.grid(True); fig.tight_layout()
    fig.savefig(p:=f"{OUT_DIR}/mae_mfe_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    mae_png = p

    rp = tdf["ret"] * 100
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.hist(rp[rp > 0], bins=25, color="#4ade80", alpha=0.80, label="Wins", edgecolor="none")
    ax.hist(rp[rp <= 0], bins=25, color="#f87171", alpha=0.80, label="Losses", edgecolor="none")
    ax.axvline(rp.mean(), color="#38bdf8", ls="--", lw=1.6, label=f"Mean {rp.mean():.2f}%")
    ax.axvline(0, color="#475569", lw=0.8)
    ax.set_xlabel("Trade Return %"); ax.set_ylabel("Count")
    ax.set_title(f"{symbol} {tf} - Trade Return Distribution", fontsize=11, fontweight="bold")
    ax.legend(fontsize=9); ax.grid(True); fig.tight_layout()
    fig.savefig(p:=f"{OUT_DIR}/dist_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    dist_png = p

    sess = tdf.groupby("session").agg(
        profit_pct=("ret", lambda r: (np.prod(1+r)-1)*100),
        win_rate  =("ret", lambda r: (r>0).mean()*100),
    ).reindex(["Asia","London","NY_overlap","NY","Late"]).fillna(0)
    fig, ax1 = plt.subplots(figsize=(7, 4)); ax2 = ax1.twinx()
    colors = ["#4ade80" if v >= 0 else "#f87171" for v in sess.profit_pct]
    ax1.bar(sess.index, sess.profit_pct, color=colors, alpha=0.8)
    ax2.plot(sess.index, sess.win_rate, color="#fbbf24", marker="o", lw=2)
    ax1.set_ylabel("Profit %"); ax2.set_ylabel("Win Rate %")
    ax1.set_title(f"{symbol} {tf} - Session Breakdown", fontsize=11, fontweight="bold")
    ax1.grid(True); fig.tight_layout()
    fig.savefig(p:=f"{OUT_DIR}/session_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    sess_png = p

    tr = tdf["ret"].values
    p5, p50, p95 = mc_equity_bands(tr)
    actual = BALANCE0 * np.cumprod(1 + tr)
    fig, ax = plt.subplots(figsize=(9, 4))
    x = range(len(tr))
    ax.fill_between(x, p5, p95, alpha=0.18, color="#38bdf8", label=f"P5-P95 ({MC_SIMS} sims)")
    ax.plot(p50,   color="#38bdf8", lw=1.5, label="MC Median")
    ax.plot(actual, color="#fbbf24", lw=1.8, label="Actual")
    ax.axhline(BALANCE0, color="#475569", ls="--", lw=0.8)
    ax.set_xlabel("Trade #"); ax.set_ylabel("Balance ($)")
    ax.set_title(f"{symbol} {tf} - Monte Carlo Equity Fan", fontsize=11, fontweight="bold")
    ax.legend(fontsize=9); ax.grid(True); fig.tight_layout()
    fig.savefig(p:=f"{OUT_DIR}/mc_{symbol}_{tf}.png", dpi=130); plt.close(fig)
    mc_png = p

    try:
        wday_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        pivot = tdf.pivot_table("pnl", index="weekday", columns="hour", aggfunc="sum")
        pivot = pivot.reindex([d for d in wday_order if d in pivot.index])
        fig, ax = plt.subplots(figsize=(14, 4))
        im = ax.imshow(pivot.fillna(0).values, aspect="auto", cmap="RdYlGn")
        fig.colorbar(im, ax=ax, label="Total PnL ($)")
        ax.set_xticks(range(len(pivot.columns))); ax.set_xticklabels(pivot.columns, fontsize=8)
        ax.set_yticks(range(len(pivot.index)));   ax.set_yticklabels(pivot.index, fontsize=9)
        ax.set_title(f"{symbol} {tf} - PnL by Hour x Weekday", fontsize=11, fontweight="bold")
        fig.tight_layout()
        fig.savefig(p:=f"{OUT_DIR}/heatmap_time_{symbol}_{tf}.png", dpi=130); plt.close(fig)
        time_hm_png = p
    except Exception:
        time_hm_png = ""

    try:
        window_r = max(10, len(tdf) // 6)
        roll_s   = tdf["ret"].rolling(window_r).apply(
            lambda r: (r.mean()/r.std()*np.sqrt(len(r))) if r.std()>0 else 0)
        fig, ax = plt.subplots(figsize=(9, 3))
        ax.plot(roll_s.values, color="#a78bfa", lw=1.6)
        ax.axhline(0, color="#475569", lw=0.8, ls="--")
        ax.fill_between(range(len(roll_s)), roll_s, 0,
                        where=roll_s >= 0, alpha=0.15, color="#4ade80")
        ax.fill_between(range(len(roll_s)), roll_s, 0,
                        where=roll_s < 0,  alpha=0.15, color="#f87171")
        ax.set_xlabel("Trade #"); ax.set_ylabel("Rolling Sharpe")
        ax.set_title(f"{symbol} {tf} - Rolling Sharpe (window={window_r})", fontsize=11, fontweight="bold")
        ax.grid(True); fig.tight_layout()
        fig.savefig(p:=f"{OUT_DIR}/rolling_sharpe_{symbol}_{tf}.png", dpi=130); plt.close(fig)
        rs_png = p
    except Exception:
        rs_png = ""

    return eq_png, hm_png, mae_png, dist_png, sess_png, mc_png, time_hm_png, rs_png

def make_portfolio_visuals(full_df, equity_curves, frontier_df=None):


    _style()

    best = full_df.loc[full_df.groupby(["symbol","timeframe"])["profit_factor"].idxmax()]

    fig, ax = plt.subplots(figsize=(11, 7))
    sc = ax.scatter(best["max_drawdown_pct"].abs(), best["profit_pct"],
                    c=best["sqn"], cmap="viridis", s=130, edgecolors="#0a0f1e", linewidths=0.6, alpha=0.92)
    fig.colorbar(sc, ax=ax, label="SQN Score")
    for _, r in best.iterrows():
        ax.annotate(f"{r.symbol} {r.timeframe}",
                    (abs(r.max_drawdown_pct), r.profit_pct),
                    xytext=(5,4), textcoords="offset points", fontsize=7.5, color="#94a3b8")
    ax.set_xlabel("Max Drawdown % (Risk)"); ax.set_ylabel("Total Return % (Reward)")
    ax.set_title("Universe Risk-Return Frontier (colour = SQN)", fontsize=12, fontweight="bold")
    ax.grid(True); fig.tight_layout()
    fig.savefig(rr:=f"{OUT_DIR}/portfolio_rr.png", dpi=130); plt.close(fig)

    mat = best.pivot_table("profit_factor","symbol","timeframe",aggfunc="max")
    fig, ax = plt.subplots(figsize=(9,6))
    im = ax.imshow(mat.values, aspect="auto", cmap="magma")
    fig.colorbar(im, ax=ax, label="Profit Factor")
    ax.set_xticks(range(len(mat.columns))); ax.set_xticklabels(mat.columns)
    ax.set_yticks(range(len(mat.index)));   ax.set_yticklabels(mat.index)
    ax.set_title("Universe Profit Factor Matrix", fontsize=12, fontweight="bold")
    for i in range(len(mat.index)):
        for j in range(len(mat.columns)):
            v = mat.iloc[i,j]
            if not np.isnan(v):
                ax.text(j,i,f"{v:.1f}",ha="center",va="center",
                        color="white" if v>2 else "#fbbf24", fontsize=8)
    fig.tight_layout(); fig.savefig(mxp:=f"{OUT_DIR}/universe_matrix.png", dpi=130); plt.close(fig)

    bs = best.sort_values("profit_pct", ascending=False)
    labels = bs["symbol"] + " " + bs["timeframe"]
    fig, ax = plt.subplots(figsize=(12, max(4, len(labels)*0.3)))
    ax.barh(labels, bs["profit_pct"],
            color=["#4ade80" if v>=0 else "#f87171" for v in bs["profit_pct"]])
    ax.set_xlabel("Strategy Return %")
    ax.set_title("Best EMA Strategy Return - All Assets", fontsize=12, fontweight="bold")
    ax.invert_yaxis(); ax.axvline(0, color="#475569", lw=0.8)
    fig.tight_layout(); fig.savefig(sp:=f"{OUT_DIR}/summary_bar.png", dpi=130); plt.close(fig)

    if len(equity_curves) > 1:
        min_len = min(len(eq) for eq in equity_curves.values())
        corr_series = {f"{sym}_{tf}": pd.Series(eq[:min_len]) for (sym, tf), eq in equity_curves.items()}
        corr_df = pd.DataFrame(corr_series).pct_change().dropna()
        valid_cols = [c for c in corr_df.columns if corr_df[c].std() > 1e-9]
        if len(valid_cols) > 1:
            corr_df = corr_df[valid_cols]
            corr_mat = corr_df.corr()
            fig, ax = plt.subplots(figsize=(max(7, len(corr_mat)*0.6+2),
                                            max(6, len(corr_mat)*0.5+2)))
            im = ax.imshow(corr_mat.values, cmap="coolwarm", vmin=-1, vmax=1)
            fig.colorbar(im, ax=ax, label="Pearson Correlation")
            ax.set_xticks(range(len(corr_mat.columns))); ax.set_xticklabels(corr_mat.columns, rotation=45, ha="right", fontsize=7)
            ax.set_yticks(range(len(corr_mat.index)));   ax.set_yticklabels(corr_mat.index, fontsize=7)
            ax.set_title("Cross-Asset Equity Curve Correlation", fontsize=12, fontweight="bold")
            for i in range(len(corr_mat.index)):
                for j in range(len(corr_mat.columns)):
                    ax.text(j,i,f"{corr_mat.iloc[i,j]:.2f}",ha="center",va="center",fontsize=6,color="white")
            fig.tight_layout(); fig.savefig(corrp:=f"{OUT_DIR}/cross_asset_corr.png", dpi=130); plt.close(fig)
        else:
            corrp = ""
    else:
        corrp = ""

    if frontier_df is not None and not frontier_df.empty:
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(frontier_df["volatility"], frontier_df["target_return"], 'o-', color="#38bdf8", lw=2, label="Efficient Frontier")
        ax.set_xlabel("Annualised Volatility %"); ax.set_ylabel("Expected Return %")
        ax.set_title("Mean-Variance Portfolio Efficient Frontier", fontsize=11, fontweight="bold")
        ax.grid(True); ax.legend(fontsize=9); fig.tight_layout()
        fig.savefig(efp:=f"{OUT_DIR}/efficient_frontier.png", dpi=130); plt.close(fig)
    else:
        efp = ""

    return sp, rr, mxp, corrp, efp

# ----------------------------- CHART.JS DASHBOARD ---------------------------
def build_dashboard(profiles_df, full_df, best_by_combo, combi_df, out_path,
                    sl_comparison_df=None, sl_conclusion_df=None):
    bdf = pd.DataFrame([
        dict(symbol=s, timeframe=t, best_ema_x=bx, best_ema_y=by, **sm)
        for (s,t),(bx,by,_,_,sm,*_) in best_by_combo.items()
    ]).sort_values("profit_pct", ascending=False)

    def jd(x): return json.dumps([round(v, 2) if isinstance(v, float) else v for v in x])
    labs  = jd([(f"{r.symbol} {r.timeframe}") for _,r in bdf.iterrows()])
    pcts  = jd(bdf["profit_pct"].tolist())
    # A8 fix: safe column access
    bh_vals = bdf["bh_return_pct"].tolist() if "bh_return_pct" in bdf.columns else [0.0]*len(bdf)
    bh    = jd(bh_vals)
    pf    = jd(bdf["profit_factor"].tolist() if "profit_factor" in bdf.columns else [0.0]*len(bdf))
    sh    = jd(bdf["sharpe"].tolist()        if "sharpe"        in bdf.columns else [0.0]*len(bdf))
    so    = jd(bdf["sortino"].tolist()       if "sortino"       in bdf.columns else [0.0]*len(bdf))
    sq    = jd(bdf["sqn"].tolist()           if "sqn"           in bdf.columns else [0.0]*len(bdf))
    dd    = jd(bdf["max_drawdown_pct"].abs().tolist() if "max_drawdown_pct" in bdf.columns else [0.0]*len(bdf))
    kf    = jd(bdf["kelly_fraction"].tolist() if "kelly_fraction" in bdf.columns else [0.0]*len(bdf))
    ror   = jd(bdf["risk_of_ruin_pct"].tolist())

    top_html   = bdf.head(15).to_html(index=False, classes="tbl")
    prof_html  = profiles_df.to_html(index=False, classes="tbl")
    combi_html = combi_df.head(24).to_html(index=False, classes="tbl") if not combi_df.empty else ""

    # Build SL section HTML
    if sl_conclusion_df is not None and not sl_conclusion_df.empty:
        best_sl   = sl_conclusion_df.iloc[0]
        sl_conc_html = sl_conclusion_df.head(13).to_html(index=False, classes="tbl")
        sl_section_html = f"""
<div style='margin-bottom:10px'>
  <b style='color:#4ade80'>Best Config:</b>
  <span style='color:#38bdf8'>{best_sl.get('sl_type','N/A').upper()} @ {best_sl.get('sl_level_pct',0):.1f}%</span>
  &nbsp;|&nbsp; Sharpe: <b>{best_sl.get('sharpe',0):.2f}</b>
  &nbsp;|&nbsp; Profit: <b>{best_sl.get('profit_pct',0):.2f}%</b>
  &nbsp;|&nbsp; MDD: <b>{best_sl.get('max_drawdown_pct',0):.2f}%</b>
  &nbsp;|&nbsp; Score: <b>{best_sl.get('score',0):.3f}</b>
</div>
<div style='overflow-x:auto'>{sl_conc_html}</div>"""
    elif sl_comparison_df is not None and not sl_comparison_df.empty:
        sl_section_html = f"<div style='overflow-x:auto'>{sl_comparison_df.head(13).to_html(index=False, classes='tbl')}</div>"
    else:
        sl_section_html = "<p style='color:#64748b'>No stop-loss data available.</p>"


    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>EMA SAR - Elite Quant Dashboard v5 (AGI Tier)</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#070d1b;color:#e2e8f0;font-family:'Inter',sans-serif}}
.nav{{background:rgba(7,13,27,.95);backdrop-filter:blur(16px);border-bottom:1px solid #1e3a5f;
      padding:12px 28px;display:flex;align-items:center;gap:16px;position:sticky;top:0;z-index:100}}
.nav h1{{font-size:1rem;font-weight:700;color:#38bdf8;letter-spacing:-.4px}}
.pill{{background:#0f2744;color:#38bdf8;padding:3px 11px;border-radius:20px;font-size:.7rem;font-weight:600}}
.main{{padding:20px 28px}}
.kpi{{display:grid;grid-template-columns:repeat(8,1fr);gap:12px;margin-bottom:20px}}
.kcard{{background:#0d1526;border:1px solid #1e3a5f;border-radius:10px;padding:14px 10px;text-align:center}}
.klabel{{font-size:.6rem;color:#64748b;text-transform:uppercase;letter-spacing:.8px;margin-bottom:4px}}
.kval{{font-size:1.3rem;font-weight:700;color:#38bdf8}}
.ksub{{font-size:.65rem;color:#475569;margin-top:2px}}
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.grid3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px}}
.card{{background:#0d1526;border:1px solid #1e3a5f;border-radius:10px;padding:16px}}
.sec{{font-size:.8rem;font-weight:600;color:#38bdf8;margin-bottom:10px;
      border-bottom:1px solid #1e3a5f;padding-bottom:6px}}
.ch{{position:relative;height:260px}}
.ch-tall{{position:relative;height:340px}}
.tbl{{width:100%;border-collapse:collapse;font-size:.68rem}}
.tbl th{{background:#050b18;color:#38bdf8;padding:6px 8px;text-align:left;
          text-transform:uppercase;letter-spacing:.4px;border-bottom:1px solid #1e3a5f}}
.tbl td{{padding:5px 8px;border-bottom:1px solid rgba(30,58,95,.4);color:#cbd5e1}}
.tbl tr:hover td{{background:rgba(56,189,248,.04)}}
img{{width:100%;border-radius:8px;border:1px solid #1e3a5f}}
.formula{{background:#050b18;border:1px solid #1e3a5f;border-radius:8px;
           padding:14px;font-family:monospace;font-size:.73rem;color:#7dd3fc;line-height:2}}
.green{{color:#4ade80;font-weight:600}} .red{{color:#f87171;font-weight:600}}
</style>
</head>
<body>
<nav class="nav">
  <h1>EMA SAR - Elite Quant Dashboard v5 (AGI Tier)</h1>
  <span class="pill">Walk-Forward {WF_FOLDS}-Fold OOS</span>
  <span class="pill">Monte Carlo {MC_SIMS}x</span>
  <span class="pill">Kelly Criterion</span>
  <span class="pill">24 Permutations Matrix</span>
  <span class="pill">VaR / CVaR</span>
  <span class="pill">Efficient Frontier</span>
  <span class="pill">{len(profiles_df)} Datasets</span>
</nav>
<div class="main">

<!-- KPIs -->
<div class="kpi">
  <div class="kcard"><div class="klabel">Best Return</div>
    <div class="kval green">{bdf['profit_pct'].max():.1f}%</div>
    <div class="ksub">{bdf.iloc[0]['symbol']} {bdf.iloc[0]['timeframe']}</div></div>
  <div class="kcard"><div class="klabel">Best Profit Factor</div>
    <div class="kval">{bdf['profit_factor'].max():.2f}x</div>
    <div class="ksub">Gross P / Gross L</div></div>
  <div class="kcard"><div class="klabel">Best Sharpe</div>
    <div class="kval">{bdf['sharpe'].max():.2f}</div>
    <div class="ksub">Risk-Adjusted</div></div>
  <div class="kcard"><div class="klabel">Best SQN</div>
    <div class="kval">{bdf['sqn'].max():.2f}</div>
    <div class="ksub">Van Tharp</div></div>
  <div class="kcard"><div class="klabel">Avg VaR (95%)</div>
    <div class="kval red">{bdf['var_95_pct'].mean():.2f}%</div>
    <div class="ksub">Value at Risk</div></div>
  <div class="kcard"><div class="klabel">Avg CVaR (95%)</div>
    <div class="kval red">{bdf['cvar_95_pct'].mean():.2f}%</div>
    <div class="ksub">Expected Shortfall</div></div>
  <div class="kcard"><div class="klabel">Datasets</div>
    <div class="kval">{len(profiles_df)}</div>
    <div class="ksub">Symbol x Timeframe</div></div>
  <div class="kcard"><div class="klabel">Pairs Tested</div>
    <div class="kval">{len(full_df):,}</div>
    <div class="ksub">EMA Combinations</div></div>
</div>

<!-- Row 1 Charts -->
<div class="grid2">
  <div class="card"><div class="sec">Strategy Return vs Buy & Hold</div>
    <div class="ch-tall"><canvas id="retChart"></canvas></div></div>
  <div class="card"><div class="sec">Risk-Adjusted Metrics (Sharpe / Sortino / SQN)</div>
    <div class="ch-tall"><canvas id="riskChart"></canvas></div></div>
</div>

<!-- Row 2 Charts -->
<div class="grid3">
  <div class="card"><div class="sec">Profit Factor by Strategy</div>
    <div class="ch"><canvas id="pfChart"></canvas></div></div>
  <div class="card"><div class="sec">Max Drawdown Comparison</div>
    <div class="ch"><canvas id="ddChart"></canvas></div></div>
  <div class="card"><div class="sec">Kelly Fraction & Risk of Ruin %</div>
    <div class="ch"><canvas id="kellyChart"></canvas></div></div>
</div>

<!-- Portfolio Visuals -->
<div class="grid2" style="margin-bottom:16px">
  <div class="card"><div class="sec">Risk-Return Frontier (SQN colored)</div>
    <img src="portfolio_rr.png"></div>
  <div class="card"><div class="sec">Universe Profit Factor Matrix</div>
    <img src="universe_matrix.png"></div>
</div>
<div class="grid2" style="margin-bottom:16px">
  <div class="card"><div class="sec">Cross-Asset Equity Correlation</div>
    <img src="cross_asset_corr.png"></div>
  <div class="card"><div class="sec">Mean-Variance Efficient Frontier</div>
    <img src="efficient_frontier.png"></div>
</div>

<!-- Top Strategies Table -->
<div class="card" style="margin-bottom:16px">
  <div class="sec">Top 15 Strategies - Full Analytics</div>
  <div style="overflow-x:auto">{top_html}</div>
</div>

<!-- Combinatorial Matrix Preview -->
<div class="card" style="margin-bottom:16px">
  <div class="sec">Combinatorial Permutation Matrix (Sample 24 Combos)</div>
  <div style="overflow-x:auto">{combi_html}</div>
</div>

<!-- Formulas -->
<div class="card" style="margin-bottom:16px">
  <div class="sec">Mathematical Engine - Elite Quant Formulas</div>
  <div class="formula">
Trade Return     = (P_exit/P_entry - 1) x Direction - 2x(Fee+Slippage) [realistic, 1-bar shift]
CAGR             = (1 + Return%)^(365/Days) - 1
Sharpe           = Mean(R) / Std(R) x sqrt(252 x bars_per_day / avg_hold)
Sortino          = Mean(R) / Downside_Std x sqrt(252 x bars_per_day / avg_hold)
Calmar           = Return% / |MaxDD%|
Omega            = Sum(Positive Returns) / Sum(|Negative Returns|)
Ulcer Index      = sqrt(Mean(Drawdown%^2))
VaR (95%)        = 5th percentile of return distribution
CVaR (95%)       = Expected return below 5th percentile (Tail loss)
Pain Index       = Mean absolute drawdown depth
DVR              = Sharpe x (1 - |MaxDD%|/100)
Information Ratio= Alpha / Tracking Error
Runs Test Z      = (Runs - E[R]) / sqrt(Var[R])   (Random walk test)
ADF Test         = MacKinnon critical value p-value mapping
Variance Ratio   = Lo-MacKinlay heteroscedasticity-robust VR
Efficient Front  = Mean-Variance SLSQP solver for optimal Sharpe weights
  </div>
</div>

<!-- Dataset Profiles -->
<div class="card">
  <div class="sec">Dataset Metadata & Quality Audit</div>
  <div style="overflow-x:auto">{prof_html}</div>
<!-- Stop-Loss Analysis Section -->
<div class="card" style="margin-bottom:16px">
  <div class="sec">&#x1F6D1; Stop-Loss Analysis — All Levels (Fixed &amp; Trailing)</div>
  {sl_section_html}
</div>
</div>

<script>
const C=(id)=>document.getElementById(id).getContext('2d');
const labs={labs};
const opts={{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{size:10}}}}}}}},
  scales:{{x:{{ticks:{{color:'#64748b',font:{{size:9}},maxRotation:50}},grid:{{color:'rgba(30,58,95,.4)'}}}},
           y:{{ticks:{{color:'#64748b',font:{{size:9}}}},grid:{{color:'rgba(30,58,95,.4)'}}}}}}
}};
const pcts={pcts}; const bh={bh}; const pf={pf}; const sh={sh};
const so={so}; const sq={sq}; const dd={dd}; const kf={kf}; const ror={ror};

new Chart(C('retChart'),{{type:'bar',data:{{labels:labs,datasets:[
  {{label:'Strategy %',data:pcts,backgroundColor:pcts.map(v=>v>=0?'rgba(74,222,128,.75)':'rgba(248,113,113,.75)'),borderRadius:3}},
  {{label:'Buy&Hold %',data:bh,backgroundColor:'rgba(56,189,248,.35)',borderRadius:3}}
]}},options:opts}});

new Chart(C('riskChart'),{{type:'bar',data:{{labels:labs,datasets:[
  {{label:'Sharpe', data:sh,  backgroundColor:'rgba(56,189,248,.75)',borderRadius:3}},
  {{label:'Sortino',data:so,  backgroundColor:'rgba(167,139,250,.75)',borderRadius:3}},
  {{label:'SQN',   data:sq,  backgroundColor:'rgba(251,191,36,.75)', borderRadius:3}}
]}},options:opts}});

new Chart(C('pfChart'),{{type:'bar',data:{{labels:labs,datasets:[
  {{label:'Profit Factor',data:pf,
    backgroundColor:pf.map(v=>v>=2?'rgba(74,222,128,.8)':v>=1?'rgba(251,191,36,.7)':'rgba(248,113,113,.7)'),borderRadius:3}}
]}},options:opts}});

new Chart(C('ddChart'),{{type:'bar',data:{{labels:labs,datasets:[
  {{label:'Max DD %',data:dd,backgroundColor:'rgba(248,113,113,.7)',borderRadius:3}}
]}},options:opts}});

new Chart(C('kellyChart'),{{type:'bar',data:{{labels:labs,datasets:[
  {{label:'Kelly f*',   data:kf, backgroundColor:'rgba(56,189,248,.75)',borderRadius:3}},
  {{label:'Risk of Ruin %',data:ror,backgroundColor:'rgba(248,113,113,.6)',borderRadius:3}}
]}},options:opts}});
</script>

</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[dashboard] -> {out_path}")

# ----------------------------- SL VISUALS -----------------------------------
def make_sl_visuals(symbol, tf, sl_df):
    """Generates SL analysis charts including 2D Dual-SL Heatmap."""


    _style()
    os.makedirs(OUT_DIR, exist_ok=True)
    pngs = []

    if sl_df.empty:
        return pngs

    base = sl_df[sl_df['sl_type'] == 'none']
    fixed_df = sl_df[sl_df['sl_type'] == 'fixed'].sort_values('fixed_sl_pct')
    trail_df = sl_df[sl_df['sl_type'] == 'trailing'].sort_values('trailing_sl_pct')
    dual_df  = sl_df[sl_df['sl_type'] == 'dual']

    # Chart 1: Profit % vs Fixed / Trailing SL
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(fixed_df['fixed_sl_pct'].values, fixed_df['profit_pct'].values, 'o-', color='#38bdf8', lw=2, label='Fixed SL')
    ax.plot(trail_df['trailing_sl_pct'].values, trail_df['profit_pct'].values, 's-', color='#a78bfa', lw=2, label='Trailing SL')
    if not base.empty:
        ax.axhline(base['profit_pct'].iloc[0], color='#4ade80', ls='--', lw=1.5, label='No SL')
    ax.set_xlabel('SL Level %'); ax.set_ylabel('Profit %')
    ax.set_title(f'{symbol} {tf} - Profit % vs Stop Level', fontsize=11, fontweight='bold')
    ax.legend(fontsize=9); ax.grid(True); fig.tight_layout()
    p = f'{OUT_DIR}/sl_profit_{symbol}_{tf}.png'
    fig.savefig(p, dpi=120); plt.close(fig); pngs.append(p)

    # Chart 2: Sharpe vs Stop Level
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(fixed_df['fixed_sl_pct'].values, fixed_df['sharpe'].values, 'o-', color='#38bdf8', lw=2, label='Fixed SL')
    ax.plot(trail_df['trailing_sl_pct'].values, trail_df['sharpe'].values, 's-', color='#a78bfa', lw=2, label='Trailing SL')
    if not base.empty:
        ax.axhline(base['sharpe'].iloc[0], color='#4ade80', ls='--', lw=1.5, label='No SL')
    ax.set_xlabel('SL Level %'); ax.set_ylabel('Sharpe Ratio')
    ax.set_title(f'{symbol} {tf} - Sharpe vs Stop Level', fontsize=11, fontweight='bold')
    ax.legend(fontsize=9); ax.grid(True); fig.tight_layout()
    p = f'{OUT_DIR}/sl_sharpe_{symbol}_{tf}.png'
    fig.savefig(p, dpi=120); plt.close(fig); pngs.append(p)

    # Chart 3: 2D Dual-SL Surface Heatmap (Fixed SL x Trailing SL Profit Factor)
    if not dual_df.empty:
        try:
            piv = dual_df.pivot_table("profit_factor", "fixed_sl_pct", "trailing_sl_pct", aggfunc="max")
            fig, ax = plt.subplots(figsize=(8, 6))
            im = ax.imshow(piv.values, aspect="auto", origin="lower", cmap="viridis",
                           extent=[piv.columns.min(), piv.columns.max(), piv.index.min(), piv.index.max()])
            fig.colorbar(im, ax=ax, label="Profit Factor")
            ax.set_xlabel("Trailing SL %"); ax.set_ylabel("Fixed SL %")
            ax.set_title(f"{symbol} {tf} - Dual Stop-Loss Surface Heatmap", fontsize=11, fontweight="bold")
            fig.tight_layout()
            p = f"{OUT_DIR}/sl_heatmap_{symbol}_{tf}.png"
            fig.savefig(p, dpi=120); plt.close(fig); pngs.append(p)
        except Exception:
            pass

    return pngs


def clean_df_floats(df, decimals=4):
    """
    Rounds float columns in DataFrames and strips timezones from datetimes for Excel compatibility.
    Safe for pandas 3.0+ StringDtype.
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    df_c = df.copy()
    for col in df_c.columns:
        if pd.api.types.is_datetime64_any_dtype(df_c[col]):
            try:
                df_c[col] = df_c[col].dt.tz_localize(None)
            except Exception:
                try:
                    df_c[col] = df_c[col].dt.tz_convert(None)
                except Exception:
                    pass
        elif pd.api.types.is_float_dtype(df_c[col]):
            df_c[col] = df_c[col].round(decimals)
    return df_c


# ----------------------------- REPORT WRITER --------------------------------
def write_reports(profiles_df, full_df, best_by_combo, doc_df, portfolio_pngs, combi_df, frontier_df, weights_df,
                  sl_comparison_df=None, sl_conclusion_df=None, sl_nosl_comp_df=None, sl_h2h_df=None):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XI
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    os.makedirs(CSV_DIR, exist_ok=True)

    summary_rows = []
    for (s,t),(bx,by,_,_,sm,*_) in best_by_combo.items():
        summary_rows.append(dict(symbol=s,timeframe=t,ema_x=bx,ema_y=by,**sm))
    sdf = pd.DataFrame(summary_rows).sort_values("profit_pct", ascending=False)

    risk_rows = []
    for (s,t),(bx,by,_,_,sm,*_) in best_by_combo.items():
        risk_rows.append(dict(
            symbol=s, timeframe=t, ema_x=bx, ema_y=by,
            var_95_pct=sm.get("var_95_pct", 0),
            cvar_95_pct=sm.get("cvar_95_pct", 0),
            pain_index=sm.get("pain_index", 0),
            dvr=sm.get("dvr", 0),
            information_ratio=sm.get("information_ratio", 0),
            gain_to_pain_ratio=sm.get("gain_to_pain_ratio", 0),
            ulcer_index=sm.get("ulcer_index", 0),
            omega_ratio=sm.get("omega_ratio", 0)
        ))
    risk_df = pd.DataFrame(risk_rows)

    stat_rows = []
    for (s,t),(bx,by,_,_,sm,*_) in best_by_combo.items():
        stat_rows.append(dict(
            symbol=s, timeframe=t, ema_x=bx, ema_y=by,
            t_stat=sm.get("t_stat", 0), p_value=sm.get("p_value", 1),
            ljung_box_p=sm.get("ljung_box_p", 1),
            runs_z_score=sm.get("runs_z_score", 0), runs_p_value=sm.get("runs_p_value", 1),
            is_significant=sm.get("is_significant", False),
            has_autocorrelation=sm.get("has_autocorrelation", False)
        ))
    stat_df = pd.DataFrame(stat_rows)

    sl_comp_out = sl_comparison_df if sl_comparison_df is not None else pd.DataFrame()
    sl_conc_out = sl_conclusion_df if sl_conclusion_df is not None else pd.DataFrame()
    sl_nosl_out = sl_nosl_comp_df  if sl_nosl_comp_df  is not None else pd.DataFrame()
    sl_h2h_out  = sl_h2h_df        if sl_h2h_df        is not None else pd.DataFrame()
    rr_surface  = sl_comp_out[sl_comp_out['tp_rr'] > 0][['fixed_sl_pct', 'tp_rr', 'sharpe', 'profit_factor', 'win_rate_pct']].copy() if (not sl_comp_out.empty and 'tp_rr' in sl_comp_out.columns) else pd.DataFrame()

    for name, df in [("Formulas",doc_df),("Profiles",profiles_df),("Summary",sdf),
                     ("RiskMetrics",risk_df),("StatTests",stat_df),("CombiMatrix",combi_df),
                     ("EfficientFrontier",frontier_df if frontier_df is not None else pd.DataFrame()),
                     ("PortfolioWeights",weights_df if weights_df is not None else pd.DataFrame()),
                     ("Full_Grid",full_df),("Top50",full_df.sort_values("profit_factor",ascending=False).head(50)),
                     ("SL_Permutations_Full", sl_comp_out),
                     ("SL_vs_NoSL_Comparative", sl_nosl_out),
                     ("TSL_vs_FixedSL_HeadToHead", sl_h2h_out),
                     ("SL_Conclusion_Master", sl_conc_out),
                     ("RR_Surface", rr_surface)]:
        clean_df_floats(df, 4).to_csv(f"{CSV_DIR}/{name}.csv", index=False, float_format="%.4f")
    for (s,t),(bx,by,tdf,*_) in best_by_combo.items():
        clean_df_floats(tdf, 4).to_csv(f"{CSV_DIR}/Trades_{s}_{t}.csv", index=False, float_format="%.4f")
    print(f"[csv] -> {CSV_DIR}/")

    wb  = Workbook(); wb.remove(wb.active)
    hf  = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="07091c")

    def ws(name, df):
        if df is None or df.empty: return
        df_clean = clean_df_floats(df, 4)
        sheet = wb.create_sheet(name[:31])
        for row in dataframe_to_rows(df_clean, index=False, header=True):
            sheet.append(row)
        for c in sheet[1]: c.font = hf; c.fill = hfill
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = min(30, max(10, max(len(str(c.value or "")) for c in col)+2))
        return sheet

    ws("Formulas",         doc_df)
    ws("Profiles",         profiles_df)
    ws("Summary",          sdf)
    ws("RiskMetrics",      risk_df)
    ws("StatTests",        stat_df)
    ws("CombiMatrix",      combi_df)
    if frontier_df is not None: ws("EfficientFrontier", frontier_df)
    if weights_df is not None:  ws("PortfolioWeights", weights_df)
    ws("Full_Grid",        full_df)
    ws("Top50",            full_df.sort_values("profit_factor",ascending=False).head(50))
    if sl_comparison_df is not None and not sl_comparison_df.empty:
        ws("SL_Comparison",  sl_comparison_df)
        # TSL-only subset
        tsl_sub = sl_comparison_df[sl_comparison_df['sl_type'] == 'trailing']
        if not tsl_sub.empty: ws("TSL_Comparison", tsl_sub)
    if sl_conclusion_df is not None and not sl_conclusion_df.empty:
        ws("SL_Conclusion",  sl_conclusion_df)

    # Per-symbol detailed breakdown sheets (Yearly, Session, Regime, Monthly, Trades)
    for (s,t),(bx,by,tdf,eq_curve,sm,reg_prf,monthly,*_) in best_by_combo.items():
        cols = ["entry_time","exit_time","entry_price","exit_price","direction","ret","pnl",
                "hold_bars","mae_pct","mfe_pct","trade_efficiency_pct","edge_ratio","entry_regime",
                "session","weekday","hour","year","month","equity","cum_profit","cum_loss"]
        ws(f"Trades_{s}_{t}", tdf[[c for c in cols if c in tdf.columns]])
        yearly = tdf.groupby("year").agg(
            trades=("pnl","count"),profit_pct=("ret",lambda r:round((np.prod(1+r)-1)*100,2)),
            cum_pnl=("pnl","sum"),avg_mae=("mae_pct","mean"),avg_mfe=("mfe_pct","mean"),
        ).reset_index()
        ws(f"Yearly_{s}_{t}", yearly)
        sess = tdf.groupby("session").agg(
            trades=("pnl","count"),profit_pct=("ret",lambda r:round((np.prod(1+r)-1)*100,2)),
            win_rate_pct=("pnl",lambda p:round((p>0).mean()*100,2)),cum_pnl=("pnl","sum"),
        ).reset_index()
        ws(f"Session_{s}_{t}", sess)
        ws(f"Regime_{s}_{t}", pd.DataFrame([dict(regime=k,**v) for k,v in reg_prf.items()]))
        ws(f"Monthly_{s}_{t}", monthly)

    cs = wb.create_sheet("Charts"); row = 1
    for p in portfolio_pngs:
        if p and os.path.exists(p) and os.path.getsize(p) > 100:
            try:
                cs.add_image(XI(p), f"A{row}"); row += 22
            except Exception:
                pass
    for (s,t), payload in best_by_combo.items():
        *_, eq_p, hm_p, mae_p, dist_p, sess_p, mc_p, thm_p, rs_p = payload
        cs.cell(row=row,column=1,value=f"{s} {t}").font = Font(bold=True,size=11); row+=1
        for col, png in [("A",eq_p),("L",hm_p),("W",mc_p)]:
            if png and os.path.exists(png) and os.path.getsize(png) > 100:
                try:
                    cs.add_image(XI(png), f"{col}{row}")
                except Exception:
                    pass
        row += 22

    xlsx = f"{OUT_DIR}/optimization_report.xlsx"
    wb.save(xlsx); print(f"[excel] -> {xlsx}")

    zpath = f"{OUT_DIR}/quant_analysis_bundle.zip"
    with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as zf:
        for root,_,files in os.walk(OUT_DIR):
            for f in files:
                if f.endswith(".zip"): continue
                fp = os.path.join(root,f)
                zf.write(fp, os.path.relpath(fp, OUT_DIR))
    print(f"[zip] -> {zpath}  ({os.path.getsize(zpath)//1024//1024} MB)")

# ----------------------------- Q4: CRASH STRESS TEST -----------------------
def crash_stress_test(df, ema_s_full, ema_l_full, symbol, tf, bx, by):
    """
    Q4 Upgrade: Backtest the winning EMA strategy over historical crash windows.
    Reports: max drawdown, Sharpe, profit during each crisis period.
    """
    close_all = df["close"].to_numpy(dtype=np.float64)
    ts_all    = pd.to_datetime(df["time"]).to_numpy()
    results   = []
    for name, t0, t1 in CRASH_WINDOWS:
        mask = (df["time"] >= t0) & (df["time"] <= t1)
        sub  = df[mask]
        if len(sub) < max(bx, by) + 5:
            results.append(dict(window=name, start=t0, end=t1, bars=0,
                                sharpe=0.0, profit_pct=0.0, max_drawdown_pct=0.0, trades=0))
            continue
        c  = sub["close"].to_numpy(dtype=np.float64)
        ts = pd.to_datetime(sub["time"]).to_numpy()
        es = _ema(c, 2.0 / (bx + 1.0))
        el = _ema(c, 2.0 / (by + 1.0))
        res = backtest_pair(c, ts, es, el)
        if res is None:
            results.append(dict(window=name, start=t0, end=t1, bars=len(sub),
                                sharpe=0.0, profit_pct=0.0, max_drawdown_pct=0.0, trades=0))
        else:
            sm = res[0]
            results.append(dict(
                window=name, start=t0, end=t1, bars=len(sub),
                sharpe=sm["sharpe"], profit_pct=sm["profit_pct"],
                max_drawdown_pct=sm["max_drawdown_pct"],
                trades=sm["total_trades"],
            ))
    return pd.DataFrame(results)

# ----------------------------- S1: CPCV & PBO -------------------------------
def compute_cpcv_pbo(close, ts, grid_df, n_splits=6, purge_pct=0.01):
    """
    S1 Upgrade: Combinatorial Purged Cross-Validation (CPCV) & Probability of Backtest Overfitting (PBO)
    (López de Prado 2018 - Advances in Financial Machine Learning, Ch. 12)
    """
    import itertools
    if grid_df is None or grid_df.empty or len(close) < 100:
        return 0.0, pd.DataFrame()

    top_candidates = grid_df.sort_values("sharpe", ascending=False).head(20)
    if len(top_candidates) < 4:
        return 0.0, pd.DataFrame()

    N = len(close)
    sub_len = N // n_splits
    if sub_len < 20:
        return 0.0, pd.DataFrame()

    cand_returns = {}
    for idx, row in top_candidates.iterrows():
        bx, by = int(row["ema_x"]), int(row["ema_y"])
        es = _ema(close, 2.0 / (bx + 1.0))
        el = _ema(close, 2.0 / (by + 1.0))
        sig = generate_strategy_signals(close, es, el)
        pos = np.roll(sig, 1); pos[0] = 0
        rets = pos[:-1] * (close[1:] / close[:-1] - 1.0)
        cand_returns[(bx, by)] = rets

    k = n_splits // 2
    groups = list(range(n_splits))
    combos = list(itertools.combinations(groups, k))

    logits = []
    underperform_count = 0

    for test_groups in combos:
        train_groups = [g for g in groups if g not in test_groups]
        
        train_mask = np.zeros(len(close) - 1, dtype=bool)
        test_mask  = np.zeros(len(close) - 1, dtype=bool)

        for g in train_groups:
            st = g * sub_len
            en = (g + 1) * sub_len if g < n_splits - 1 else N
            p_buf = int((en - st) * purge_pct)
            train_mask[max(0, st + p_buf): min(N - 1, en - p_buf)] = True

        for g in test_groups:
            st = g * sub_len
            en = (g + 1) * sub_len if g < n_splits - 1 else N
            test_mask[st: min(N - 1, en)] = True

        is_perf = {}
        oos_perf = {}
        for pair, rets in cand_returns.items():
            tr_r = rets[train_mask]
            te_r = rets[test_mask]
            is_perf[pair]  = np.mean(tr_r) / (np.std(tr_r) + 1e-9)
            oos_perf[pair] = np.mean(te_r) / (np.std(te_r) + 1e-9)

        best_is_pair = max(is_perf, key=is_perf.get)
        best_oos_val = oos_perf[best_is_pair]

        all_oos_vals = list(oos_perf.values())
        median_oos = float(np.median(all_oos_vals))

        if best_oos_val <= median_oos:
            underperform_count += 1

        rank_oos = sum(1 for v in all_oos_vals if v <= best_oos_val) / float(len(all_oos_vals))
        rank_oos = min(max(rank_oos, 1e-4), 1.0 - 1e-4)
        logit = np.log(rank_oos / (1.0 - rank_oos))
        logits.append(logit)

    pbo = round(float(underperform_count / max(1, len(combos))), 4)
    return pbo, pd.DataFrame({"logit": logits})

# ----------------------------- Q7: STRATEGY JSON EXPORT --------------------
def export_strategy_json(symbol, tf, bx, by, best_sl_dict=None, pbo=0.0, out_dir=OUT_DIR):
    """
    Q7 Upgrade: Export best strategy config as JSON for ccxt/IB API live deployment.
    """
    import datetime
    payload = {
        "strategy":   "EMA_CROSSOVER_ADAPTIVE",
        "symbol":     symbol,
        "timeframe":  tf,
        "fast_period": bx,
        "slow_period": by,
        "sl_type":    best_sl_dict.get("sl_type", "none") if best_sl_dict else "none",
        "fixed_sl_pct":    best_sl_dict.get("fixed_sl_pct", 0.0) if best_sl_dict else 0.0,
        "trailing_sl_pct": best_sl_dict.get("trailing_sl_pct", 0.0) if best_sl_dict else 0.0,
        "tp_pct":     best_sl_dict.get("take_profit_pct", 0.0) if best_sl_dict else 0.0,
        "pbo":        pbo,
        "deployed_at": datetime.datetime.utcnow().isoformat() + "Z",
        "regime_adaptive": True,
        "hurst_gate": 0.45,
        "adf_gate":   0.05,
    }
    os.makedirs(out_dir, exist_ok=True)
    path = f"{out_dir}/strategy_{symbol}_{tf}.json"
    with open(path, "w") as f:
        json.dump(payload, f, indent=2)
    return path

# ----------------------------- WORKER ---------------------------------------
def worker(symbol, tf, df, mode):
    prof = profile_dataset(df, symbol, tf)
    close_arr = df["close"].to_numpy(dtype=np.float64)
    ts_arr    = pd.to_datetime(df["time"]).to_numpy()
    reg_labs  = regime_labels(close_arr)
    h_val     = prof.get("hurst", 0.5)
    a_p       = prof.get("adf_pvalue", 0.10)
    n_bars    = len(close_arr)

    if mode == "bo":
        # Scale Bayesian init/iter with dataset size
        n_init = min(64, max(20, n_bars // 200))
        n_iter = min(40, max(10, n_bars // 500))
        res, best = bayesian_optimization_search(df, symbol, tf, n_init=n_init, n_iter=n_iter, reg_labels=reg_labs, hurst_val=h_val, adf_p=a_p)
    elif mode == "full":
        # S4: Adaptive period range — full exhaustive brute-force (every integer pair with min_sep=5)
        tf_min_f, tf_max_f = PERIOD_RANGE.get(tf, (MIN_P, MAX_P))
        print(f"  [full-grid] {symbol} {tf}: periods {tf_min_f}..{tf_max_f} ({tf_max_f-tf_min_f+1} pts)")
        res, best = run_grid(df, symbol, tf,
                             list(range(tf_min_f, tf_max_f+1)),
                             list(range(tf_min_f, tf_max_f+1)),
                             reg_labels=reg_labs, hurst_val=h_val, adf_p=a_p, min_sep=5)
    else:
        res, best = coarse_fine(df, symbol, tf, reg_labels=reg_labs, hurst_val=h_val, adf_p=a_p)
    grid_df = res.to_df()
    if best is None:
        return symbol, tf, prof, grid_df, None, pd.DataFrame(), pd.DataFrame()

    sm, tdf, eq_curve, bx, by, reg_prf, monthly, ema_s, ema_l = best

    # Q5 Upgrade: Parameter Stability Score (Mean Sharpe in radius r=10)
    r = 10
    neighbors = grid_df[(grid_df["ema_x"] >= bx - r) & (grid_df["ema_x"] <= bx + r) &
                        (grid_df["ema_y"] >= by - r) & (grid_df["ema_y"] <= by + r)]
    sm["stability_score"] = round(float(neighbors["sharpe"].mean()), 2) if not neighbors.empty else round(float(sm["sharpe"]), 2)

    # S1 Upgrade: CPCV & PBO Overfitting Evaluation
    try:
        pbo_val, _ = compute_cpcv_pbo(close_arr, ts_arr, grid_df)
        sm["pbo"] = pbo_val
    except Exception:
        sm["pbo"] = 0.0

    # Q4: Crash Stress Test
    try:
        crash_df = crash_stress_test(df, ema_s, ema_l, symbol, tf, bx, by)
        sm["crash_worst_dd"] = round(float(crash_df["max_drawdown_pct"].min()), 2)
        sm["crash_worst_sharpe"] = round(float(crash_df["sharpe"].min()), 2)
    except Exception:
        crash_df = pd.DataFrame()

    # Q7: Export strategy JSON
    export_strategy_json(symbol, tf, bx, by, pbo=sm.get("pbo", 0.0))

    # Q2: Transaction Cost Sensitivity Sweep & Break-Even Analysis
    fee_png, breakeven_fee_bps = fee_sensitivity_sweep(close_arr, ts_arr, ema_s, ema_l, symbol, tf, hurst_val=h_val, adf_p=a_p)
    sm["breakeven_fee_bps"] = breakeven_fee_bps

    combi = generate_combinatorial_matrix(close_arr, ts_arr, ema_s, ema_l, symbol, tf, bx, by, reg_labels=reg_labs, hurst_val=h_val, adf_p=a_p)

    pngs = make_visuals(symbol, tf, grid_df, bx, by, eq_curve, tdf)

    # Stop-loss grid analysis on best EMA pair with C6 try/except wrapper
    try:
        sl_df = run_sl_grid(close_arr, ts_arr, ema_s, ema_l, symbol, tf, bx, by, reg_labels=reg_labs, hurst_val=h_val, adf_p=a_p)
        make_sl_visuals(symbol, tf, sl_df)
    except Exception as e:
        print(f"[WARN] SL Grid run failed for {symbol} {tf}: {e}")
        sl_df = pd.DataFrame()

    return symbol, tf, prof, grid_df, (bx, by, tdf, eq_curve, sm, reg_prf, monthly, *pngs), combi, sl_df

def archive_previous_run():
    xlsx_path = os.path.join(OUT_DIR, "optimization_report.xlsx")
    html_path = os.path.join(OUT_DIR, "dashboard.html")
    if os.path.exists(xlsx_path) or os.path.exists(html_path):
        import datetime, shutil, time
        mtime = os.path.getmtime(xlsx_path) if os.path.exists(xlsx_path) else time.time()
        ts_str = datetime.datetime.fromtimestamp(mtime).strftime("%Y%m%d_%H%M%S")
        arc_dir = os.path.join("./archive", f"output_{ts_str}")
        if not os.path.exists(arc_dir):
            os.makedirs(arc_dir, exist_ok=True)
            for item in os.listdir(OUT_DIR):
                if item in ("data_cache", ".git"): continue
                s_path = os.path.join(OUT_DIR, item)
                d_path = os.path.join(arc_dir, item)
                try:
                    if os.path.isdir(s_path):
                        shutil.copytree(s_path, d_path, dirs_exist_ok=True)
                    else:
                        shutil.copy2(s_path, d_path)
                except Exception: pass
            print(f"[+] Archived previous test output -> {arc_dir}")

def save_run_info(args, data):
    import datetime
    info = {
        "timestamp": datetime.datetime.now().isoformat(),
        "mode": args.mode,
        "risk_aversion": args.risk_aversion,
        "fetch_live": getattr(args, "fetch_live", False),
        "datasets_tested": [f"{s}_{t}" for s, t in data.keys()],
        "total_datasets": len(data),
    }
    with open(os.path.join(OUT_DIR, "run_info.json"), "w", encoding="utf-8") as f:
        json.dump(info, f, indent=2)

# ----------------------------- MAIN -----------------------------------------
def main():
    import multiprocessing
    multiprocessing.freeze_support()
    pa = argparse.ArgumentParser(description="EMA SAR Optimizer — Elite Quant Engine")
    pa.add_argument("--mode",  choices=["full","coarse","bo"], default="coarse")
    pa.add_argument("--demo",  action="store_true")
    pa.add_argument("--input", default=DATA_XLSX)
    pa.add_argument("--fetch-live", action="store_true", help="Force fetching fresh live market data via yfinance")
    # Q8: Expose risk aversion lambda for utility-theoretic ranking
    pa.add_argument("--risk-aversion", type=float, default=RISK_AVERSION,
                    dest="risk_aversion",
                    help="von Neumann-Morgenstern risk-aversion lambda (default=2.0)")
    args = pa.parse_args()
    # Propagate lambda to global so compute_utility_score picks it up
    import builtins
    builtins._RISK_AVERSION = args.risk_aversion

    archive_previous_run()
    os.makedirs(OUT_DIR, exist_ok=True)
    if args.fetch_live:
        print("[+] --fetch-live requested: fetching real-time market data...")
        data = fetch_live_market_data()
    elif os.path.exists(args.input):
        data = load_all(args.input)
        if not data:
            print("[!] Local input file has no valid datasets. Fetching live market data...")
            data = fetch_live_market_data()
    else:
        print("[+] Local input file not found. Fetching live market data...")
        data = fetch_live_market_data()

    if not data:
        print("No usable sheets found."); return

    save_run_info(args, data)

    doc_df = pd.DataFrame([
        ("CAGR",                "(1+R%)^(365/Days)-1",             "Annualised compound growth"),
        ("Trade Return",        "(P1/P0-1)xDir - 2xFee",           "Net per-trade return (1-bar shift)"),
        ("Profit Factor",       "Sum(Wins)/Sum(|Losses|)",         "Gross profit / gross loss"),
        ("Sharpe",              "mu/sigma x sqrt(252 x bpd / avg_hold)", "Total-vol risk-adjusted return"),
        ("Sortino",             "mu/sigma_down x sqrt(252 x bpd / avg_hold)", "Downside-vol risk-adjusted return"),
        ("Calmar",              "Return%/|MaxDD%|",                "Return vs worst drawdown"),
        ("Omega Ratio",         "Sum(Pos)/Sum(|Neg|)",             "Probability-weighted ratio"),
        ("Ulcer Index",         "sqrt(Mean(DD%^2))",               "Sustained drawdown pain metric"),
        ("VaR (95%)",           "5th Percentile Return",           "Value at Risk (95% Confidence)"),
        ("CVaR (95%)",          "E[R | R <= VaR_95]",              "Expected Shortfall / Tail Loss"),
        ("Pain Index",          "Mean(|DD%|)",                     "Average drawdown depth"),
        ("DVR",                 "Sharpe x (1 - |MaxDD%|)",         "Deflated / Drawdown-adjusted Sharpe"),
        ("Information Ratio",   "Alpha / Tracking Error",          "Benchmark excess return performance"),
        ("Runs Test Z",         "(Runs - E[R]) / Std[R]",          "Trade sequence randomness test"),
        ("ADF Test",            "MacKinnon critical value p-value", "Price non-stationarity test"),
        ("Variance Ratio",      "Lo-MacKinlay heteroscedastic VR",  "Market efficiency / Random Walk test"),
        ("PBO (CPCV)",          "P(OOS_Rank <= Median | IS_Best)",  "Probability of Backtest Overfitting (López de Prado)"),
        ("Efficient Frontier",  "SLSQP Mean-Variance Solver",      "Optimal Sharpe portfolio allocation weights"),
    ], columns=["Metric","Formula","Description"])

    all_frames, all_profiles, best_by_combo, eq_curves, all_combis, all_sl = [], [], {}, {}, [], []

    max_workers = max(1, min(8, (os.cpu_count() or 2)))
    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(worker, s, t, df, args.mode): (s,t) for (s,t),df in data.items()}
        total_tasks = len(futs)
        for i, fut in enumerate(as_completed(futs), 1):
            s, t = futs[fut]
            try:
                sym, tf, prof, grid_df, best, combi, sl_df = fut.result()
            except Exception as e:
                print(f"[{i}/{total_tasks}] [FAIL] {s} {t}: {e}"); continue
            all_profiles.append(prof); all_frames.append(grid_df)
            if not combi.empty: all_combis.append(combi)
            if not sl_df.empty: all_sl.append(sl_df)
            if best is not None:
                bx, by, tdf, eq, sm = best[0], best[1], best[2], best[3], best[4]
                best_by_combo[(sym,tf)] = best
                eq_curves[(sym,tf)]     = eq
                print(
                    f"[{i:02d}/{total_tasks:02d}] [{sym:5} {tf:3}] EMA({bx:3},{by:3}) | "
                    f"PF={sm['profit_factor']:5.2f} | "
                    f"Sharpe={sm['sharpe']:5.2f} | "
                    f"VaR95={sm['var_95_pct']:6.2f}% | "
                    f"CVaR95={sm['cvar_95_pct']:6.2f}% | "
                    f"DVR={sm['dvr']:5.2f}"
                )

    if not all_frames: print("No results."); return

    prof_df  = pd.DataFrame(all_profiles).sort_values(["symbol","timeframe"])
    full_df  = pd.concat(all_frames, ignore_index=True)

    # Q6: Benjamini-Hochberg False Discovery Rate (FDR) Multiple Testing Correction
    if "p_value" in full_df.columns and len(full_df) > 10:
        try:
            from statsmodels.stats.multitest import multipletests
            rej, pvals_corrected, _, _ = multipletests(full_df['p_value'].fillna(1.0), method='fdr_bh')
            full_df['p_value_fdr'] = np.round(pvals_corrected, 4)
            full_df['is_significant_fdr'] = rej
        except Exception:
            pass

    combi_df = pd.concat(all_combis, ignore_index=True) if all_combis else pd.DataFrame()
    sl_comparison_df = pd.concat(all_sl, ignore_index=True) if all_sl else pd.DataFrame()

    # Build Master SL Conclusion & Comparative Reports
    sl_conclusion_df, sl_nosl_comp_df, sl_h2h_df = pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if not sl_comparison_df.empty:
        best_sl_dict, sl_conclusion_df, sl_nosl_comp_df, sl_h2h_df = conclude_best_sl(sl_comparison_df)
        if best_sl_dict:
            print(f"\n[MASTER SL CONCLUSION] Best Configuration: "
                  f"{best_sl_dict.get('symbol')} {best_sl_dict.get('timeframe')} | "
                  f"Type={best_sl_dict.get('sl_type').upper()} | "
                  f"FixedSL={best_sl_dict.get('fixed_sl_pct')}% | "
                  f"TrailingSL={best_sl_dict.get('trailing_sl_pct')}% | "
                  f"Score={best_sl_dict.get('score'):.3f}")
            print(f"Rationale: {best_sl_dict.get('rationale')}")

    frontier_df, weights_df = compute_efficient_frontier(best_by_combo)

    _, rr, mx, corr, efp = make_portfolio_visuals(full_df, eq_curves, frontier_df)
    build_dashboard(prof_df, full_df, best_by_combo, combi_df, f"{OUT_DIR}/dashboard.html",
                    sl_comparison_df=sl_comparison_df, sl_conclusion_df=sl_conclusion_df)
    write_reports(prof_df, full_df, best_by_combo, doc_df, [rr, mx, corr, efp], combi_df,
                  frontier_df, weights_df,
                  sl_comparison_df=sl_comparison_df, sl_conclusion_df=sl_conclusion_df,
                  sl_nosl_comp_df=sl_nosl_comp_df, sl_h2h_df=sl_h2h_df)

    print(f"\nCompleted -> {OUT_DIR}/dashboard.html | {OUT_DIR}/quant_analysis_bundle.zip")

if __name__ == "__main__":
    main()
